"""
Verificador de Comprovantes — Currículo Lattes
===============================================
Lê o lattes.xml, verifica PDFs comprovantes e gera relatório Excel por
candidato + planilha de ranking geral.

Estrutura de pastas:
    candidatos/
        NomeCandidato/
            lattes.xml
            1.1.pdf, 1.2.pdf ...   (artigos)
            2.1.pdf ...             (trabalhos completos)
            3.1.pdf ...             (resumos)
            4.1.pdf ...             (livros)
            5.1.pdf ...             (capítulos)
            6.1.pdf ...             (orientações)
            7.1.pdf ...             (atuação profissional — docente ensino superior)
            8.1.pdf ...             (atuação profissional — docente ensino básico)
            9.1.pdf ...             (atuação profissional — não docente)
            10.1.pdf ...            (projetos de pesquisa)
            11.1.pdf ...            (projetos de extensão)

Arquivos de apoio (mesma pasta que verificador.py):
    criterios.xlsx   — pontuações editáveis
    qualis.xlsx               — tabela ISSN → Estrato (aba "Educação Física")
"""

import xml.etree.ElementTree as ET
import pdfplumber
import unicodedata
import re
import math
import sys
from pathlib import Path
from datetime import date, datetime

# scipy é opcional: usado para a atribuição ótima PDF↔item do Lattes (algoritmo
# húngaro). Sem ele, cai para uma atribuição gulosa por score decrescente
# (_atribuicao_otima) — pior que o ótimo em casos raros de empate, mas ainda
# assim melhor que "cada PDF pega guloso o seu melhor item isoladamente".
try:
    import numpy as np
    from scipy.optimize import linear_sum_assignment
    _HAS_SCIPY = True
except ImportError:
    _HAS_SCIPY = False
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Caminhos padrão dos arquivos de apoio ─────────────────────────────────────

_DIR = Path(__file__).parent
CRITERIOS_XLSX = _DIR / "criterios.xlsx"
QUALIS_XLSX    = _DIR / "qualis.xlsx"
CANDIDATOS_DIR = _DIR / "candidatos"


# ── Normalização de texto ─────────────────────────────────────────────────────

def normalizar(texto: str) -> str:
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto.lower()).strip()


def similaridade(a: str, b: str) -> float:
    stop = {"a", "o", "e", "de", "do", "da", "em", "no", "na",
            "the", "of", "in", "and", "for", "to", "with"}
    # Os dois lados precisam tokenizar do MESMO jeito (\w+, via regex) — usar
    # .split() de um lado e regex do outro faz palavra com vírgula/dois-pontos
    # grudados ("intake,") ou hífen ("cross-sectional") nunca bater contra o
    # texto do PDF, mesmo quando a palavra está lá, só que sem a pontuação
    # (o PDF é tokenizado por \w+, que já separa nesses caracteres).
    palavras = [p for p in re.findall(r"\w+", normalizar(a)) if len(p) > 3 and p not in stop]
    if not palavras:
        return 0.0
    tokens_b = set(re.findall(r"\w+", normalizar(b)))
    return sum(1 for p in palavras if p in tokens_b) / len(palavras)


# Termos genéricos demais para diferenciar UMA instituição de outra — se
# entrassem na conta de similaridade, duas instituições completamente
# diferentes do mesmo estado/área poderiam "bater" só por coincidência
# (ex.: "Faculdade de Ensino de Minas Gerais" x "Universidade do Estado de
# Minas Gerais" compartilham "ensino"/"minas"/"gerais" sem serem a mesma).
_STOP_INSTITUICAO = {
    "universidade", "faculdade", "instituto", "centro", "escola", "colegio", "colégio",
    "fundacao", "fundação", "sociedade", "associacao", "associação", "unidade",
    "federal", "estadual", "municipal", "particular", "publica", "pública", "privada",
    "ensino", "superior", "tecnologico", "tecnológico", "universitario", "universitário",
    "educacao", "educação", "ciencias", "ciências",
    "minas", "gerais", "paulo", "sao", "são", "janeiro", "rio", "grande", "sul", "norte",
    "distrito", "santa", "catarina", "espirito", "espírito", "santo", "bahia",
    "ceara", "ceará", "parana", "paraná", "pernambuco", "goias", "goiás",
    "amazonas", "para", "pará",
}


def similaridade_instituicao(a: str, b: str) -> float:
    """Como similaridade(), mas ignora termos genéricos demais para identificar
    uma instituição específica (tipo de instituição, "ensino/superior", nomes
    de estado). Evita falso-positivo entre instituições diferentes que só
    compartilham essas palavras comuns."""
    stop = {"a", "o", "e", "de", "do", "da", "em", "no", "na"} | _STOP_INSTITUICAO
    palavras = [p for p in re.findall(r"\w+", normalizar(a)) if len(p) > 3 and p not in stop]
    if not palavras:
        # Nome da instituição não tem NENHUMA palavra distintiva (é só termo
        # genérico + nome de estado) — não dá para confirmar com segurança
        # por bag-of-words. Falha de propósito (0.0) em vez de aprovar por
        # coincidência: força revisão manual em vez de risco de falso positivo.
        return 0.0
    tokens_b = set(re.findall(r"\w+", normalizar(b)))
    return sum(1 for p in palavras if p in tokens_b) / len(palavras)


def _ocr_pdf(caminho_pdf: Path, max_paginas: int = 3) -> str:
    """Tenta extrair texto via OCR (fallback para PDFs escaneados)."""
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except ImportError:
        return ""
    try:
        imagens = convert_from_path(str(caminho_pdf), last_page=max_paginas, dpi=200)
        return "\n".join(
            pytesseract.image_to_string(img, lang="por+eng") for img in imagens
        )
    except Exception:
        return ""


def extrair_texto_pdf(caminho_pdf: Path, max_paginas: int = 3) -> str:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = "\n".join(p.extract_text() or "" for p in pdf.pages[:max_paginas])
    except Exception as e:
        return f"__ERRO_PDF__: {e}"
    if texto.strip():
        return texto
    # PDF escaneado — tenta OCR
    ocr = _ocr_pdf(caminho_pdf, max_paginas)
    return ocr if ocr.strip() else texto


# ── Carregamento de critérios e Qualis ────────────────────────────────────────

# Número da seção (cabeçalho "N. Título...") → tipos internos candidatos
# naquela seção. Quando há mais de um candidato, _resolver_tipo desambigua
# por palavra-chave no texto da linha (Quesito + Subcategoria combinados).
_SECAO_TIPOS = {
    "1":  ["artigo"],
    "2":  ["trabalho_completo"],
    "3":  ["resumo"],
    "4":  ["livro_organizado", "livro_publicado"],
    "5":  ["capitulo"],
    "6":  ["orientacao_mestrado", "orientacao_tcc", "orientacao_ic", "orientacao_esp"],
    "7":  ["banca"],
    "8":  ["atuacao"],
    "9":  ["projeto_pesquisa"],
    "10": ["projeto_extensao"],
    "13": ["organizacao_evento"],
    "14": ["ic_manual"],
    "16": ["prod_artistica_int", "prod_artistica_nac", "prod_artistica_reg"],
}

# Tipos cuja subcategoria de pontuação é sempre fixa ("unidade") — a própria
# seção/tipo já resolve o nível (ex.: prod_artistica_int já é "internacional").
_TIPOS_UNIDADE_FIXA = frozenset({
    "prod_artistica_int", "prod_artistica_nac", "prod_artistica_reg",
    "organizacao_evento", "ic_manual",
})

# Tipos que podem ser pontuados de forma flat ("por unidade") OU, se a
# planilha voltar a diferenciar geografia, por Internacional/Nacional/Regional.
_TIPOS_GEO_OU_FLAT = frozenset({
    "trabalho_completo", "resumo", "livro_publicado", "livro_organizado", "capitulo",
})


# Palavras-chave que identificam cada tipo, testadas no texto combinado
# (Quesito + Subcategoria) — independente do número de seção da planilha.
# É a fonte primária de resolução: uma reordenação/renumeração/fusão de
# categorias na planilha (já aconteceu na prática — ver criterios.xlsx vs.
# criterios_producao.xlsx) não quebra o parser, porque ele não depende de
# "seção 7 é sempre banca". _SECAO_TIPOS vira só reserva, usada quando
# nenhuma palavra-chave bate.
#
# Ordem importa: tipos mais específicos (mais padrões exigidos, ou que
# tomariam o match de um tipo mais genérico) vêm primeiro — ex. "orientação
# de mestrado" precisa ser testado antes de "banca", "iniciação científica
# como orientador" antes de "iniciação científica como participante", e
# "livro organizado" antes de "livro publicado" (todo livro organizado
# também contém a palavra "livro"). Cada tipo é (padrões,) — todos os
# padrões precisam casar (re.search) no texto para o tipo ser aceito.
_TIPO_KEYWORDS = [
    ("orientacao_mestrado", (r"orienta", r"mestrado")),
    ("orientacao_ic",       (r"orienta", r"iniciacao cientifica")),
    ("orientacao_esp",      (r"orienta", r"especializac|aperfeicoamento")),
    ("orientacao_tcc",      (r"orienta", r"\btcc\b|graduacao|conclusao de curso")),
    ("ic_manual",           (r"iniciacao cientifica",)),
    ("banca",               (r"\bbanca",)),
    ("prod_artistica_int",  (r"producao artistica", r"internacional")),
    ("prod_artistica_nac",  (r"producao artistica", r"\bnacional\b")),
    ("prod_artistica_reg",  (r"producao artistica", r"regional|\blocal\b")),
    ("livro_organizado",    (r"\blivro", r"organiz|edic")),
    ("livro_publicado",     (r"\blivro", r"public")),
    ("capitulo",            (r"capitulo",)),
    ("trabalho_completo",   (r"trabalho completo",)),
    ("resumo",              (r"resumo", r"anais|congresso|evento")),
    ("organizacao_evento",  (r"organiz.*evento|comissao organizadora",)),
    ("atuacao",             (r"atuacao",)),
    ("projeto_pesquisa",    (r"projeto", r"pesquisa")),
    ("projeto_extensao",    (r"projeto", r"extensao")),
    ("artigo",              (r"\bartigo|periodico|qualis",)),
]


def _resolver_tipo(secao: str, texto: str) -> str | None:
    """Identifica o tipo interno de uma linha da planilha por palavra-chave no
    texto (Quesito + Subcategoria combinados) — ver _TIPO_KEYWORDS. O número
    de seção (_SECAO_TIPOS) só entra como resposta de reserva, e apenas
    quando a seção tem um único tipo candidato possível (com mais de um
    candidato e nenhuma palavra-chave batendo, não há como desambiguar com
    segurança — melhor descartar a linha do que arriscar o tipo errado)."""
    for tipo, padroes in _TIPO_KEYWORDS:
        if all(re.search(p, texto) for p in padroes):
            return tipo

    candidatos = _SECAO_TIPOS.get(secao)
    if candidatos and len(candidatos) == 1:
        return candidatos[0]
    return None


def _sub_banca(texto: str) -> str:
    """Classifica uma banca (linha da planilha ou item do Lattes, já normalizado
    ou não) em Doutorado/Mestrado/Especializacao/Graduacao por palavra-chave.
    Compartilhado entre o parser da planilha e calcular_pontos_item para que
    as duas classificações nunca divirjam."""
    t = normalizar(texto)
    if "doutorado" in t:
        return "Doutorado"
    if "mestrado" in t:
        return "Mestrado"
    if "especializac" in t or "aperfeicoamento" in t:
        return "Especializacao"
    return "Graduacao"


def _resolver_sub(tipo: str, texto: str) -> str:
    """Identifica a subcategoria (chave de pontuação) de uma linha pelo
    texto combinado (Quesito + Subcategoria), por palavra-chave de domínio."""
    if tipo == "artigo":
        m = re.search(r"\b(a1|a2|a3|a4|b1|b2|b3|b4)\b", texto)
        nivel = m.group(1).upper() if m else ("Sem Qualis" if "sem qualis" in texto else "C")
        # Quando a planilha distingue Primeiro/Demais autor (linhas
        # separadas por nível de Qualis, cada uma com pontuação própria —
        # caso do criterios.xlsx atual), a chave carrega isso junto
        # ("A1-Primeiro" / "A1-Demais"); calcular_pontos_item cai pro nível
        # "flat" (só "A1") se a planilha não fizer essa distinção.
        if "primeiro" in texto:
            return f"{nivel}-Primeiro"
        if "demais" in texto:
            return f"{nivel}-Demais"
        return nivel

    if tipo in _TIPOS_UNIDADE_FIXA:
        return "unidade"

    if tipo == "capitulo":
        # Mesmo padrão de Primeiro/Demais em artigo: quando a planilha
        # distingue autor de coautor (linhas separadas, cada uma com
        # pontuação própria — caso do criterios.xlsx atual), a chave carrega
        # isso ("Autor"/"Coautor"); calcular_pontos_item cai pro "unidade"/
        # Nacional-Internacional se a planilha não fizer essa distinção.
        if "coautor" in texto:
            return "Coautor"
        if "autor" in texto:
            return "Autor"
        # cai pro tratamento genérico abaixo (unidade/geografia)

    if tipo in _TIPOS_GEO_OU_FLAT:
        if "internacional" in texto:
            return "Internacional"
        if "nacional" in texto:
            return "Nacional"
        if "regional" in texto or "local" in texto:
            return "Regional"
        return "unidade"

    if tipo in ("orientacao_mestrado", "orientacao_tcc", "orientacao_ic", "orientacao_esp"):
        return "coorientador" if "coorientador" in texto else "principal"

    if tipo == "banca":
        return _sub_banca(texto)

    if tipo == "atuacao":
        if "basico" in texto:
            return "ensino_basico"
        if "superior" in texto:
            return "ensino_superior"
        return "nao_docencia"

    if tipo in ("projeto_pesquisa", "projeto_extensao"):
        return "coordenador" if "coordenador" in texto else "integrante"

    return "unidade"


def carregar_criterios(xlsx_path: Path = CRITERIOS_XLSX) -> tuple[dict, dict]:
    """Lê criterios.xlsx; retorna (pontos_dict, campos_config_dict).

    A seção é identificada pelo número do cabeçalho ("9. Projetos de
    pesquisa"); tipo e subcategoria são resolvidos por palavra-chave no
    texto da linha. Isso torna o parser resistente a renomeações de rótulo
    na planilha — desde que a numeração da seção e as palavras-chave de
    domínio (Qualis, Internacional/Nacional, Coordenador/Integrante,
    Doutorado/Mestrado, docência) se mantenham.
    """
    criterios: dict = {}
    campos_config: dict = {}
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb["Critérios"]
    except Exception as e:
        print(f"[AVISO] Não foi possível carregar critérios: {e}. Usando padrões.")
        return _criterios_padrao(), _campos_config_padrao()

    # Coluna "O que verificar no PDF" localizada pelo cabeçalho, não por
    # posição fixa — a planilha já teve colunas extras (Campo XML, Observação)
    # removidas/reordenadas entre revisões, o que quebraria um índice fixo.
    col_verificar = 5  # posição histórica (coluna F), usada só se o cabeçalho não for encontrado
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for i, val in enumerate(row):
            if val and "verificar" in normalizar(str(val)):
                col_verificar = i
                break

    secao_atual  = ""
    quesito_atual = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        val_a = str(row[0]).strip() if row[0] else ""
        val_b = str(row[1]).strip() if row[1] else ""
        val_c = row[2]  # pontos
        val_f = row[col_verificar] if len(row) > col_verificar else None  # O que verificar no PDF

        # Cabeçalho de seção ("N. Título..."): pode vir puro, numa linha só com
        # a coluna A preenchida (dado nas linhas seguintes — ex. "Artigos"),
        # ou já trazer o próprio dado na mesma linha, com colunas B/C também
        # preenchidas (categorias de linha única, ex. "2. Resumos publicados
        # em anais | por unidade | 0,01"). A seção precisa avançar nos dois
        # casos — senão uma categoria de linha única fica presa na seção
        # anterior e seu valor vaza para o tipo errado.
        m = re.match(r"(\d+)\.", val_a) if val_a else None
        if m:
            secao_atual = m.group(1)
        if val_a and not val_b:
            quesito_atual = val_a
            continue
        if not val_b or val_c is None:
            continue

        if val_a:
            quesito_atual = val_a

        texto = normalizar(f"{quesito_atual} {val_b}")
        tipo  = _resolver_tipo(secao_atual, texto)
        if tipo is None:
            continue
        sub = _resolver_sub(tipo, texto)

        criterios.setdefault(tipo, {})[sub] = float(val_c)
        if val_f:
            campos_config.setdefault(tipo, {})[sub] = _parse_campos_config(val_f)

    if not criterios:
        return _criterios_padrao(), _campos_config_padrao()
    return criterios, campos_config


def _criterios_padrao() -> dict:
    return {
        "artigo": {"A1": 100, "A2": 85, "A3": 70, "A4": 55,
                   "B1": 40, "B2": 25, "B3": 10, "B4": 5,
                   "C": 1, "Sem Qualis": 0.5},
        "trabalho_completo": {"unidade": 40},
        "resumo":            {"unidade": 25},
        "livro_publicado":   {"unidade": 80},
        "livro_organizado":  {"unidade": 50},
        "capitulo":          {"unidade": 30},
        "orientacao_tcc": {"principal": 10, "coorientador": 8},
        "orientacao_ic":  {"principal": 5},
        "orientacao_esp": {"principal": 10, "coorientador": 10},
        "banca": {"Mestrado": 20, "Graduacao": 10, "Especializacao": 15},
        "atuacao": {"ensino_superior": 6, "ensino_basico": 4, "nao_docencia": 2},
        "projeto_pesquisa": {"coordenador": 20, "integrante": 5},
        "projeto_extensao": {"coordenador": 20, "integrante": 5},
        "organizacao_evento": {"unidade": 20},
        "ic_manual":          {"unidade": 20},
        "prod_artistica_int": {"unidade": 60},
        "prod_artistica_nac": {"unidade": 40},
        "prod_artistica_reg": {"unidade": 20},
    }


def _campos_config_padrao() -> dict:
    """Campos por (tipo, subcategoria) — mesma granularidade de _criterios_padrao().
    Usado apenas quando a planilha não pôde ser carregada; _resolver_campos()
    cai para qualquer subcategoria disponível caso a exata não exista aqui."""
    return {
        "artigo":            {"A1": ["titulo", "doi", "autor"]},
        "trabalho_completo": {"unidade": ["titulo", "evento"]},
        "resumo":            {"unidade": ["titulo", "evento"]},
        "livro_publicado":   {"unidade": ["titulo", "isbn"]},
        "livro_organizado":  {"unidade": ["titulo", "isbn"]},
        "capitulo":          {"unidade": ["titulo", "isbn"]},
        "orientacao_mestrado": {"principal": ["titulo", "orientando"]},
        "orientacao_tcc":    {"principal": ["titulo", "orientando"]},
        "orientacao_ic":     {"principal": ["titulo", "orientando"]},
        "orientacao_esp":    {"principal": ["titulo", "orientando"]},
        "banca": {
            "Mestrado":       ["titulo", "orientando"],
            "Graduacao":      ["titulo", "candidato"],
            "Especializacao": ["titulo", "orientando"],
            "Doutorado":      ["titulo", "orientando"],
        },

        "organizacao_evento": {"unidade": ["titulo", "evento"]},
        "ic_manual":         {"unidade": ["apenas_pdf"]},
        "prod_artistica_int": {"unidade": ["titulo"]},
        "prod_artistica_nac": {"unidade": ["titulo"]},
        "prod_artistica_reg": {"unidade": ["titulo"]},
        "projeto_pesquisa":  {"coordenador": ["titulo", "autor"]},
        "projeto_extensao":  {"coordenador": ["titulo", "autor"]},
    }


def _resolver_campos(campos_por_sub: dict | None, sub: str) -> list[str]:
    """Escolhe os campos de verificação de uma subcategoria específica
    (ex.: banca 'Graduacao' vs 'Mestrado'). Cai para qualquer subcategoria
    disponível no mesmo tipo se a exata não estiver configurada, e para
    ["titulo"] se o tipo não tiver nenhuma configuração."""
    if not campos_por_sub:
        return ["titulo"]
    if sub in campos_por_sub:
        return campos_por_sub[sub]
    return next(iter(campos_por_sub.values()))


def _norm_periodico(nome: str) -> str:
    """Normaliza nome de periódico pra comparação: minúsculo, sem acento, sem
    pontuação, espaços colapsados ("Physical & Occupational..." e "Physical
    and Occupational..." caem na mesma chave)."""
    t = normalizar(nome)
    t = re.sub(r"[^\w\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


# Termos genéricos demais pra identificar UM periódico específico — se
# entrassem na comparação, "Revista Cena" (como o pesquisador escreveu no
# Lattes) nunca bateria contra "Cena. UFRGS" (como consta no Qualis) por
# igualdade de string, e um filtro fraco arriscaria "Revista" sozinho
# batendo com qualquer coisa.
_STOP_PERIODICO = {"revista", "jornal", "journal", "boletim", "the", "of", "in",
                    "a", "o", "e", "de", "do", "da", "para", "em", "no", "na"}


def _qualis_por_nome_aproximado(periodico: str, titulos: list[tuple[str, str]]) -> str | None:
    """Busca aproximada de Qualis por nome de periódico — só usada quando
    ISSN e nome normalizado exato falham (ver calcular_pontos_item).
    Cobre variação de como o nome foi cadastrado no Lattes vs. como consta
    no Qualis (ex.: "Revista Cena" x "Cena. UFRGS" — mesmo periódico, ISSN
    diferente por ser edição impressa/online, nome com prefixo genérico
    diferente).

    Exige que TODAS as palavras distintivas do nome do Lattes (depois de
    tirar termos genéricos como "revista") apareçam no título do Qualis —
    não é maioria/limiar parcial como similaridade(), de propósito: nome de
    periódico é curto, então um match parcial tem risco alto de pegar o
    periódico errado. Retorna None (sem match) se não sobrar nenhuma
    palavra distintiva pra comparar."""
    palavras = [p for p in re.findall(r"\w+", _norm_periodico(periodico))
                if len(p) > 3 and p not in _STOP_PERIODICO]
    if not palavras:
        return None
    for titulo_norm, estrato in titulos:
        tokens = set(titulo_norm.split())
        if all(p in tokens for p in palavras):
            return estrato
    return None


def carregar_qualis(xlsx_path: Path = QUALIS_XLSX) -> dict:
    """Retorna dict Estrato indexado por ISSN_sem_traço E por nome
    normalizado do periódico (aba 'Educação Física').

    O nome entra como chave de fallback: o ISSN cadastrado no Lattes às
    vezes não é o mesmo que consta no Qualis (ex.: pesquisador cadastrou o
    ISSN impresso/antigo do periódico, Qualis usa o ISSN online/atual) —
    sem o fallback por nome, o artigo cai em "Sem Qualis" por um problema de
    cadastro, não por realmente não ter Qualis. Ver calcular_pontos_item.
    """
    qualis: dict = {}
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb["Educação Física"]
    except Exception as e:
        print(f"[AVISO] Não foi possível carregar Qualis: {e}")
        return qualis

    # Linha de cabeçalho (tem "ISSN") e coluna "Estrato" localizadas pelo
    # nome, não por índice fixo — a planilha já mudou de layout entre
    # revisões (coluna "Área de Avaliação" inserida entre "Título" e
    # "Estrato", deslocando o índice), o que quebraria um `row[2]` fixo
    # silenciosamente (leria outra coluna sem dar erro).
    linha_cabecalho, col_estrato = 0, 2
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        if any(val and normalizar(str(val)) == "issn" for val in row):
            linha_cabecalho = i
            for k, val in enumerate(row):
                if val and "estrato" in normalizar(str(val)):
                    col_estrato = k
            break

    if not linha_cabecalho:
        print("[AVISO] Cabeçalho ('ISSN'/'Estrato') não encontrado no Qualis — planilha em formato inesperado.")
        wb.close()
        return qualis

    titulos: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=linha_cabecalho + 1, values_only=True):
        issn    = str(row[0]).replace("-", "").strip() if row[0] else ""
        titulo  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        estrato = str(row[col_estrato]).strip() if len(row) > col_estrato and row[col_estrato] else ""
        if issn and estrato:
            qualis[issn] = estrato
        if titulo and estrato:
            qualis.setdefault(_norm_periodico(titulo), estrato)
            titulos.append((_norm_periodico(titulo), estrato))
    # Lista de (nome, estrato) pra busca aproximada (_qualis_por_nome_aproximado)
    # — cai numa chave reservada no mesmo dict pra não mudar a assinatura da
    # função em todo lugar que já usa `qualis` (chave impossível de colidir
    # com ISSN ou nome normalizado real).
    qualis["__titulos__"] = titulos
    wb.close()
    return qualis


# ── Parsers do XML ────────────────────────────────────────────────────────────

def parse_xml(caminho_xml: Path) -> ET.Element:
    """Lê e parseia o XML do Lattes. O Lattes exporta o currículo como um
    .zip contendo o XML de verdade com nome numérico (tipo
    "6855775711261227.xml") — às vezes esse .zip é salvo/renomeado direto
    como "lattes.xml" sem ser extraído primeiro. Detecta isso pela
    assinatura do ZIP ("PK" no início do arquivo) e extrai o XML de dentro
    em memória, sem precisar que o usuário descompacte manualmente."""
    dados = caminho_xml.read_bytes()
    if dados[:2] == b"PK":
        import zipfile, io
        with zipfile.ZipFile(io.BytesIO(dados)) as z:
            nomes_xml = [n for n in z.namelist() if n.lower().endswith(".xml")]
            if not nomes_xml:
                raise ValueError(f"{caminho_xml} é um .zip mas não tem nenhum .xml dentro.")
            dados = z.read(nomes_xml[0])
    return ET.fromstring(dados.decode("iso-8859-1"))


def _achar_lattes_xml(pasta: Path) -> Path | None:
    """Localiza o XML do Lattes na pasta do candidato — nome do arquivo não
    é padronizado (varia de "lattes.xml"/"Lattes.xml" até algo como
    "Currículo Lattes - Fulano.xml"). Tenta primeiro "lattes.xml" (com
    variação de maiúscula — sistema de arquivos é case-sensitive no
    Linux); se não achar, usa qualquer .xml com "lattes" no nome; se ainda
    assim não achar e sobrar exatamente um .xml na pasta, assume que é
    ele — não tem outro tipo de arquivo XML esperado na pasta de um
    candidato."""
    direto = pasta / "lattes.xml"
    if direto.exists():
        return direto
    xmls = list(pasta.glob("*.xml"))
    for f in xmls:
        if f.name.lower() == "lattes.xml":
            return f
    for f in xmls:
        if "lattes" in normalizar(f.stem):
            return f
    if len(xmls) == 1:
        return xmls[0]
    return None


def _ordem_autoria(elem: ET.Element, nome_pesquisador: str) -> str:
    """'Primeiro' se o pesquisador é o autor de ORDEM-DE-AUTORIA="1" na
    produção (elemento AUTORES do Lattes), 'Demais' se aparece em outra
    posição, '' se o nome não bateu com nenhum AUTORES cadastrado (mesmo
    padrão de casamento por nome de _papel_no_projeto)."""
    partes = [p for p in normalizar(nome_pesquisador).split() if len(p) > 3]
    for autor in elem.findall("AUTORES"):
        nome_autor = normalizar(autor.get("NOME-COMPLETO-DO-AUTOR", ""))
        if sum(1 for p in partes if p in nome_autor) >= min(2, len(partes)):
            return "Primeiro" if autor.get("ORDEM-DE-AUTORIA", "") == "1" else "Demais"
    return ""


def extrair_artigos(root: ET.Element, nome_pesquisador: str = "") -> list[dict]:
    items = []
    for i, art in enumerate(root.findall(".//ARTIGO-PUBLICADO"), 1):
        db  = art.find("DADOS-BASICOS-DO-ARTIGO")
        det = art.find("DETALHAMENTO-DO-ARTIGO")
        items.append({
            "seq":      f"1.{i}",
            "titulo":   db.get("TITULO-DO-ARTIGO", "") if db is not None else "",
            "ano":      db.get("ANO-DO-ARTIGO", "")    if db is not None else "",
            "periodico": det.get("TITULO-DO-PERIODICO-OU-REVISTA", "") if det is not None else "",
            "issn":     det.get("ISSN", "")            if det is not None else "",
            "doi":      db.get("DOI", "")              if db is not None else "",
            "autoria":  _ordem_autoria(art, nome_pesquisador),
        })
    return items


def extrair_trabalhos_completos(root: ET.Element) -> list[dict]:
    items, i = [], 1
    for ev in root.findall(".//TRABALHO-EM-EVENTOS"):
        db  = ev.find("DADOS-BASICOS-DO-TRABALHO")
        det = ev.find("DETALHAMENTO-DO-TRABALHO")
        if db is not None and db.get("NATUREZA", "") == "TRABALHO_COMPLETO":
            items.append({
                "seq":          f"2.{i}",
                "titulo":       db.get("TITULO-DO-TRABALHO", ""),
                "ano":          db.get("ANO-DO-TRABALHO", ""),
                "evento":       det.get("NOME-DO-EVENTO", "")           if det is not None else "",
                "classificacao": det.get("CLASSIFICACAO-DO-EVENTO", "") if det is not None else "",
            })
            i += 1
    return items


def extrair_resumos(root: ET.Element) -> list[dict]:
    items, i = [], 1
    for ev in root.findall(".//TRABALHO-EM-EVENTOS"):
        db  = ev.find("DADOS-BASICOS-DO-TRABALHO")
        det = ev.find("DETALHAMENTO-DO-TRABALHO")
        # RESUMO_EXPANDIDO é resumo expandido — natureza distinta de RESUMO
        # no Lattes, mas o Anexo II não separa os dois ("2. Resumos
        # publicados em anais" cobre ambos). Seq usa prefixo "2" (não "3")
        # pelo mesmo motivo — é a seção que o Anexo II define pra isso, e é
        # como os candidatos de fato nomeiam os PDFs comprovantes.
        if db is not None and db.get("NATUREZA", "") in ("RESUMO", "RESUMO_EXPANDIDO"):
            items.append({
                "seq":          f"2.{i}",
                "titulo":       db.get("TITULO-DO-TRABALHO", ""),
                "ano":          db.get("ANO-DO-TRABALHO", ""),
                "evento":       det.get("NOME-DO-EVENTO", "")           if det is not None else "",
                "classificacao": det.get("CLASSIFICACAO-DO-EVENTO", "") if det is not None else "",
            })
            i += 1
    return items


def extrair_livros(root: ET.Element) -> list[dict]:
    items = []
    for i, liv in enumerate(root.findall(".//LIVRO-PUBLICADO-OU-ORGANIZADO"), 1):
        db  = liv.find("DADOS-BASICOS-DO-LIVRO")
        det = liv.find("DETALHAMENTO-DO-LIVRO")
        items.append({
            "seq":           f"4.{i}",
            "titulo":        db.get("TITULO-DO-LIVRO", "")      if db is not None else "",
            "ano":           db.get("ANO", "")                  if db is not None else "",
            "tipo":          db.get("TIPO", "")                 if db is not None else "",
            "editora":       det.get("NOME-DA-EDITORA", "")     if det is not None else "",
            "cidade_editora": det.get("CIDADE-DA-EDITORA", "")  if det is not None else "",
            "isbn":          det.get("ISBN", "")                if det is not None else "",
        })
    return items


def extrair_capitulos(root: ET.Element, nome_pesquisador: str = "") -> list[dict]:
    items = []
    for i, cap in enumerate(root.findall(".//CAPITULO-DE-LIVRO-PUBLICADO"), 1):
        db  = cap.find("DADOS-BASICOS-DO-CAPITULO")
        det = cap.find("DETALHAMENTO-DO-CAPITULO")
        # "Primeiro"/"Demais" de _ordem_autoria (mesma nomenclatura de
        # artigo) vira "Autor"/"Coautor" — nomenclatura que o
        # criterios.xlsx usa pra capítulo (ver _resolver_sub).
        autoria = _ordem_autoria(cap, nome_pesquisador)
        autoria = {"Primeiro": "Autor", "Demais": "Coautor"}.get(autoria, autoria)
        items.append({
            "seq":           f"5.{i}",
            "titulo":        db.get("TITULO-DO-CAPITULO-DO-LIVRO", "") if db is not None else "",
            "ano":           db.get("ANO", "")                         if db is not None else "",
            "livro":         det.get("TITULO-DO-LIVRO", "")            if det is not None else "",
            "editora":       det.get("NOME-DA-EDITORA", "")            if det is not None else "",
            "cidade_editora": det.get("CIDADE-DA-EDITORA", "")         if det is not None else "",
            "isbn":          det.get("ISBN", "")                       if det is not None else "",
            "doi":           db.get("DOI", "")                         if db is not None else "",
            "autoria":       autoria,
        })
    return items


def extrair_orientacoes(root: ET.Element) -> list[dict]:
    items, i = [], 1
    for ori in root.findall(".//ORIENTACOES-CONCLUIDAS-PARA-MESTRADO"):
        db  = ori.find("DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO")
        det = ori.find("DETALHAMENTO-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO")
        items.append({
            "seq":            f"6.{i}",
            "titulo":         db.get("TITULO", "")                if db is not None else "",
            "ano":            db.get("ANO", "")                   if db is not None else "",
            "natureza":       "Mestrado",
            "orientando":     det.get("NOME-DO-ORIENTADO", "")    if det is not None else "",
            "tipo_orientacao": det.get("TIPO-DE-ORIENTACAO", "")  if det is not None else "",
            "instituicao":    det.get("NOME-DA-INSTITUICAO", "")  if det is not None else "",
        })
        i += 1
    nat_map = {
        "TRABALHO_DE_CONCLUSAO_DE_CURSO_GRADUACAO": "TCC Graduação",
        "INICIACAO_CIENTIFICA": "Iniciação Científica",
        "MONOGRAFIA_DE_CONCLUSAO_DE_CURSO_APERFEICOAMENTO_E_ESPECIALIZACAO": "Especialização",
    }
    for ori in root.findall(".//OUTRAS-ORIENTACOES-CONCLUIDAS"):
        db  = ori.find("DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
        det = ori.find("DETALHAMENTO-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
        nat = db.get("NATUREZA", "") if db is not None else ""
        items.append({
            "seq":            f"6.{i}",
            "titulo":         db.get("TITULO", "")              if db is not None else "",
            "ano":            db.get("ANO", "")                 if db is not None else "",
            "natureza":       nat_map.get(nat, nat),
            "orientando":     det.get("NOME-DO-ORIENTADO", "")  if det is not None else "",
            "tipo_orientacao": det.get("TIPO-DE-ORIENTACAO-CONCLUIDA", "ORIENTADOR_PRINCIPAL")
                                if det is not None else "ORIENTADOR_PRINCIPAL",
            "instituicao":    det.get("NOME-DA-INSTITUICAO", "") if det is not None else "",
        })
        i += 1
    return items


def extrair_bancas(root: ET.Element) -> list[dict]:
    items, i = [], 1
    mapa = {
        "PARTICIPACAO-EM-BANCA-DE-DOUTORADO":
            ("Doutorado",  "DADOS-BASICOS-DA-PARTICIPACAO-EM-BANCA-DE-DOUTORADO",
                           "DETALHAMENTO-DA-PARTICIPACAO-EM-BANCA-DE-DOUTORADO"),
        "PARTICIPACAO-EM-BANCA-DE-MESTRADO":
            ("Mestrado",   "DADOS-BASICOS-DA-PARTICIPACAO-EM-BANCA-DE-MESTRADO",
                           "DETALHAMENTO-DA-PARTICIPACAO-EM-BANCA-DE-MESTRADO"),
        "PARTICIPACAO-EM-BANCA-DE-GRADUACAO":
            ("Graduação",  "DADOS-BASICOS-DA-PARTICIPACAO-EM-BANCA-DE-GRADUACAO",
                           "DETALHAMENTO-DA-PARTICIPACAO-EM-BANCA-DE-GRADUACAO"),
        "PARTICIPACAO-EM-BANCA-DE-APERFEICOAMENTO-ESPECIALIZACAO":
            ("Especialização", "DADOS-BASICOS-DA-PARTICIPACAO-EM-BANCA-DE-APERFEICOAMENTO-ESPECIALIZACAO",
                               "DETALHAMENTO-DA-PARTICIPACAO-EM-BANCA-DE-APERFEICOAMENTO-ESPECIALIZACAO"),
    }
    for tag, (tipo, tag_db, tag_det) in mapa.items():
        for banca in root.findall(f".//{tag}"):
            db  = banca.find(tag_db)
            det = banca.find(tag_det)
            items.append({
                "seq":       f"6.{i}",
                "titulo":    db.get("TITULO", "")             if db is not None else "",
                "ano":       db.get("ANO", "")                if db is not None else "",
                "tipo":      tipo,
                "candidato": det.get("NOME-DO-CANDIDATO", "") if det is not None else "",
                "instituicao": det.get("NOME-INSTITUICAO", "") if det is not None else "",
            })
            i += 1
    return items


def _papel_no_projeto(proj: ET.Element, nome_pesquisador: str) -> str:
    """Retorna 'coordenador' ou 'integrante' para o pesquisador naquele projeto."""
    equipe = proj.find("EQUIPE-DO-PROJETO")
    if equipe is None:
        return "integrante"
    partes = [p for p in normalizar(nome_pesquisador).split() if len(p) > 3]
    for integ in equipe.findall("INTEGRANTES-DO-PROJETO"):
        nome_integ = normalizar(integ.get("NOME-COMPLETO", ""))
        if sum(1 for p in partes if p in nome_integ) >= min(2, len(partes)):
            return "coordenador" if integ.get("FLAG-RESPONSAVEL", "NAO") == "SIM" else "integrante"
    return "integrante"


def extrair_projetos(root: ET.Element, nome_pesquisador: str) -> tuple[list, list]:
    """Retorna (projetos_pesquisa, projetos_extensao)."""
    pesquisa, extensao = [], []
    cnt = {"PESQUISA": 0, "EXTENSAO": 0}

    for proj in root.findall(".//PROJETO-DE-PESQUISA"):
        natureza = proj.get("NATUREZA", "").upper()
        if natureza not in ("PESQUISA", "EXTENSAO"):
            continue
        nome_proj = proj.get("NOME-DO-PROJETO", "")
        ano_ini  = proj.get("ANO-INICIO", "")
        ano_fim  = proj.get("ANO-FIM", "")
        situacao = proj.get("SITUACAO", "")
        papel    = _papel_no_projeto(proj, nome_pesquisador)
        periodo  = f"{ano_ini}–{ano_fim}" if ano_fim else f"{ano_ini}–atual"

        # PROJETO-DE-PESQUISA só traz ANO-INICIO/ANO-FIM no Lattes — sem mês
        # (diferente de ATUACAO-PROFISSIONAL/VINCULOS, que tem MES-INICIO/
        # MES-FIM). Sem esse dado, assume ano cheio (janeiro a dezembro) em
        # cada extremo; projeto ainda em andamento (sem ano-fim) conta até o
        # mês atual — mesma convenção de calcular_meses().
        meses = calcular_meses("1", ano_ini, "12" if ano_fim else "", ano_fim)

        if natureza == "PESQUISA":
            cnt["PESQUISA"] += 1
            lista, seq = pesquisa, f"10.{cnt['PESQUISA']}"
        else:
            cnt["EXTENSAO"] += 1
            lista, seq = extensao, f"11.{cnt['EXTENSAO']}"

        lista.append({
            "seq":      seq,
            "titulo":   nome_proj,
            "ano":      ano_ini,
            "situacao": situacao,
            "periodo":  periodo,
            "papel":    papel,
            "meses":    meses,
        })

    return pesquisa, extensao



def _scan_pdfs_manuais(pasta: Path, sec: int, titulo: str) -> list[dict]:
    """Escaneia PDFs nomeados como {sec}.N.pdf na pasta do candidato."""
    encontrados = {}
    for pdf in list(pasta.glob(f"{sec}.*.pdf")) + list(pasta.glob(f"{sec}.*.PDF")):
        seq = pdf.stem
        if seq.split(".")[0] == str(sec):
            encontrados.setdefault(seq, pdf)
    return [{"seq": s, "titulo": titulo, "ano": "", "_pdf_path": p}
            for s, p in sorted(encontrados.items())]


_TIPOS_PROD_ARTISTICA = [
    ("ARTES-VISUAIS",
     "DADOS-BASICOS-DE-ARTES-VISUAIS",
     "DETALHAMENTO-DE-ARTES-VISUAIS"),
    ("MUSICA",
     "DADOS-BASICOS-DE-MUSICA",
     "DETALHAMENTO-DE-MUSICA"),
    ("TEATRO",
     "DADOS-BASICOS-DE-TEATRO",
     "DETALHAMENTO-DE-TEATRO"),
    ("DANCA",
     "DADOS-BASICOS-DE-DANCA",
     "DETALHAMENTO-DE-DANCA"),
    ("AUDIOVISUAL",
     "DADOS-BASICOS-DO-AUDIOVISUAL",
     "DETALHAMENTO-DO-AUDIOVISUAL"),
    ("DESIGN",
     "DADOS-BASICOS-DE-DESIGN",
     "DETALHAMENTO-DE-DESIGN"),
    ("OUTRA-PRODUCAO-ARTISTICA-E-CULTURAL",
     "DADOS-BASICOS-DE-OUTRA-PRODUCAO-ARTISTICA-E-CULTURAL",
     "DETALHAMENTO-DE-OUTRA-PRODUCAO-ARTISTICA-E-CULTURAL"),
]


def _nivel_artistica(db: ET.Element | None, det: ET.Element | None) -> tuple[str, str, int]:
    """Classifica produção artística em Internacional/Nacional/Regional via FLAG-RELEVANCIA."""
    for el in (db, det):
        if el is None:
            continue
        flag = el.get("FLAG-RELEVANCIA", "").upper()
        if "INTERN" in flag:
            return "Internacional", "prod_artistica_int", 16
        if "NACION" in flag:
            return "Nacional", "prod_artistica_nac", 17
        if "REGION" in flag or "LOCAL" in flag:
            return "Regional/Local", "prod_artistica_reg", 18
    # fallback por país de publicação
    pais = (db.get("PAIS", "") if db is not None else "").upper()
    if pais and pais not in ("BRASIL", "BRAZIL", "BR", ""):
        return "Internacional", "prod_artistica_int", 16
    return "Nacional", "prod_artistica_nac", 17


def extrair_prod_artistica_xml(root: ET.Element) -> tuple[list, list, list]:
    """Extrai produções artísticas/culturais do XML Lattes e classifica por nível."""
    internacionais, nacionais, regionais = [], [], []
    cnt = {16: 0, 17: 0, 18: 0}

    for tipo_tag, tag_db, tag_det in _TIPOS_PROD_ARTISTICA:
        for el in root.findall(f".//{tipo_tag}"):
            db  = el.find(tag_db)
            det = el.find(tag_det)
            titulo = (db.get("TITULO", "") if db is not None else "")
            ano    = (db.get("ANO", "")    if db is not None else "")
            if not titulo:
                continue
            nivel, _, secao = _nivel_artistica(db, det)
            cnt[secao] += 1
            item = {
                "seq":       f"{secao}.{cnt[secao]}",
                "titulo":    titulo,
                "ano":       ano,
                "tipo_arte": tipo_tag.replace("-", " ").title(),
                "nivel":     nivel,
            }
            if secao == 16:
                internacionais.append(item)
            elif secao == 17:
                nacionais.append(item)
            else:
                regionais.append(item)

    return internacionais, nacionais, regionais


def extrair_ic_manual(pasta_candidato: Path) -> list[dict]:
    """Gera itens de IC escaneando PDFs 14.N.pdf na pasta do candidato (sem XML)."""
    return _scan_pdfs_manuais(pasta_candidato, 14, "Iniciação Científica")


def extrair_banca_especializacao_manual(pasta_candidato: Path) -> list[dict]:
    """Gera itens de banca de especialização escaneando PDFs 15.N.pdf na pasta
    do candidato. O Lattes não tem uma tag própria para essa participação
    (só Graduação/Mestrado/Doutorado), então — como a IC — é aprovada apenas
    pela presença do PDF, sem cruzamento com o XML."""
    return _scan_pdfs_manuais(pasta_candidato, 15, "Banca de Especialização")


def extrair_organizacoes_evento(root: ET.Element) -> list[dict]:
    items = []
    for i, el in enumerate(root.findall(".//ORGANIZACAO-DE-EVENTO"), 1):
        db  = el.find("DADOS-BASICOS-DA-ORGANIZACAO-DE-EVENTO")
        det = el.find("DETALHAMENTO-DA-ORGANIZACAO-DE-EVENTO")
        items.append({
            "seq":    f"13.{i}",
            "titulo": db.get("TITULO", "")                  if db is not None else "",
            "ano":    db.get("ANO", "")                     if db is not None else "",
            "evento": det.get("INSTITUICAO-PROMOTORA", "")  if det is not None else "",
        })
    return items


# ── Atuação profissional ──────────────────────────────────────────────────────

LABEL_ATUACAO = {
    "ensino_superior": "Docência — Ensino Superior",
    "ensino_basico":   "Docência — Ensino Básico",
    "nao_docencia":    "Atividades de Não Docência",
}

_SUPERIOR_KW = [
    "universidade", "faculdade", "centro universitario", "centro universitário",
    "instituto federal", "ufrgs", "ufmg", "uemg", "fumec", "ufjf",
    "unifal", "ufla", "ufsj", "puc", "mackenzie", "ufop", "ufv",
]
_BASICO_KW = [
    "e.e.", "escola estadual", "escola municipal", "e.m.",
    "colegio", "colégio", "ensino fundamental", "ensino medio", "ensino médio",
    # Professor(a) de rede municipal/estadual no Brasil costuma ter como
    # "instituição" no Lattes a Prefeitura/Secretaria em si, não uma escola
    # nomeada — município administra diretamente educação infantil e
    # ensino fundamental (básico), não ensino superior.
    "prefeitura municipal", "prefeitura", "municipio de", "município de",
    "secretaria municipal de educacao", "secretaria estadual de educacao",
    "secretaria de educacao",
]


def classificar_atividade(instituicao: str, cargo: str) -> str:
    inst  = normalizar(instituicao)
    cargo_n = normalizar(cargo)
    if "professor" not in cargo_n and "docente" not in cargo_n:
        return "nao_docencia"
    if "ensino basico" in cargo_n or any(k in inst for k in _BASICO_KW):
        return "ensino_basico"
    if any(k in inst for k in _SUPERIOR_KW):
        return "ensino_superior"
    return "nao_docencia"


def calcular_meses(mes_ini: str, ano_ini: str, mes_fim: str, ano_fim: str) -> int:
    try:
        mi = int(mes_ini) if mes_ini else 1
        ai = int(ano_ini)
        hoje = date.today()
        mf = int(mes_fim) if mes_fim else hoje.month
        af = int(ano_fim) if ano_fim else hoje.year
        return max(0, (af - ai) * 12 + (mf - mi) + 1)
    except Exception:
        return 0


# Casa "DD/MM/AAAA <conector> DD/MM/AAAA" (conector: "a", "até"/"ate", "e",
# "-"/"–") — cobre os formatos vistos em certificados reais de projeto:
# "Período: de 01/09/2023 até 31/08/2024" e "entre 31/05/2022 e 31/03/2023".
_DATA_RANGE_RE = re.compile(
    r"(\d{1,2})/(\d{1,2})/(\d{4})\s*(?:a|at[ée]|e|-|–)\s*(\d{1,2})/(\d{1,2})/(\d{4})",
    re.IGNORECASE,
)

# Relatório institucional (ex.: "Sistema Pesquisa - Aluno" da UFRGS) rotula
# início/fim separadamente em vez de conectar as duas datas — "Início:
# 12/08/2022 Previsão de término: 29/09/2027" — sem conector simples entre
# elas, então _DATA_RANGE_RE não reconhece como par. Só usada dentro da
# busca ancorada no nome (_pares_ancorados_no_nome): fora de uma janela
# restrita a uma pessoa, esse padrão apareceria repetido pra cada membro de
# uma equipe inteira e voltaria a misturar o período de gente diferente.
_DATA_INICIO_TERMINO_RE = re.compile(
    r"in[íi]cio:?\s*(\d{1,2})/(\d{1,2})/(\d{4}).{0,60}?"
    r"(?:t[ée]rmino|conclus[ãa]o):?\s*(\d{1,2})/(\d{1,2})/(\d{4})",
    re.IGNORECASE | re.DOTALL,
)

# Carteira de Trabalho Digital / Extrato de Outros Vínculos (governo, via
# ESOCIAL) lista o contrato logo abaixo do cabeçalho "Contratos de
# trabalho" como "DD/MM/AAAA - Aberto" (ainda em aberto) ou "DD/MM/AAAA -
# DD/MM/AAAA" (encerrado) — sem isso, a busca genérica de par de data
# acaba pegando datas de outras seções do mesmo documento (ex.: "Férias
# DD/MM/AAAA a DD/MM/AAAA" em "Anotações"), que não têm nada a ver com a
# duração real do vínculo empregatício.
_CONTRATO_TRABALHO_RE = re.compile(
    r"contratos?\s+de\s+trabalho\W*"
    r"(\d{1,2})/(\d{1,2})/(\d{4})\s*-\s*(?:aberto|(\d{1,2})/(\d{1,2})/(\d{4}))",
    re.IGNORECASE,
)

# Certificado de extensão da UFRGS costuma trazer DUAS datas por atividade:
# "realizado entre X e Y" (duração NOMINAL do programa/edital inteiro — pode
# rodar o ano todo) e "tendo atuado de A até B" (participação REAL daquela
# pessoa especificamente, quase sempre mais curta). _DATA_RANGE_RE não
# distingue os dois e pega o intervalo mais amplo entre todos → sempre
# vencia o "realizado entre", inflando a participação real pro tamanho do
# programa. Esse padrão, quando presente, tem prioridade.
_DATA_ATUACAO_RE = re.compile(
    r"atuad[oa]\s+de\s+(\d{1,2})/(\d{1,2})/(\d{4})\s*(?:a|at[ée])\s*(\d{1,2})/(\d{1,2})/(\d{4})",
    re.IGNORECASE,
)


def _pares_contratos_trabalho(texto: str) -> list[tuple[str, str, str, str, str, str]]:
    """Extrai os períodos de "Contratos de trabalho" de uma Carteira de
    Trabalho Digital / Extrato de Outros Vínculos (ver
    _CONTRATO_TRABALHO_RE) — contrato "Aberto" (ainda vigente) conta até
    hoje. Documento pode listar mais de um contrato (histórico de
    empregos); retorna todos, no mesmo formato de tupla de
    _DATA_RANGE_RE.findall() pra encaixar direto no resto de
    _periodo_do_pdf."""
    pares = []
    hoje = date.today()
    for m in _CONTRATO_TRABALHO_RE.finditer(texto):
        d1, m1, a1, d2, m2, a2 = m.groups()
        try:
            date(int(a1), int(m1), int(d1))  # só valida — datas viram string de novo abaixo
        except ValueError:
            continue
        if d2 and m2 and a2:
            pares.append((d1, m1, a1, d2, m2, a2))
        else:
            pares.append((d1, m1, a1, f"{hoje.day:02d}", f"{hoje.month:02d}", str(hoje.year)))
    return pares


def _pares_ancorados_no_nome(texto: str, nome_pesquisador: str, janela: int = 130) -> list[tuple[str, str, str, str, str, str]]:
    """Procura pares de data só numa janela de texto ANCORADA em cada
    ocorrência do nome do pesquisador — usada quando o documento lista
    várias pessoas (ex.: relatório institucional de projeto com a equipe
    inteira, cada uma com seu próprio "Início"/"Previsão de término"), pra
    não misturar o período de uma pessoa com o de outra.

    Tenta o nome completo primeiro (mais específico, evita pegar
    homônimos de sobrenome); só cai pro sobrenome sozinho se o nome
    completo não aparecer em lugar nenhum do texto (documento pode
    abreviar). Busca no texto original (não normalizado) pra manter as
    posições de caractere alinhadas com o texto sendo fatiado — não dá
    pra reaproveitar `normalizar()` aqui porque ela colapsa espaços e
    mudaria o tamanho da string."""
    partes = [p for p in nome_pesquisador.split() if len(p) > 2]
    if not partes:
        return []
    for alvo in (nome_pesquisador.strip(), partes[-1]):
        if not alvo:
            continue
        pares = []
        for m in re.finditer(re.escape(alvo), texto, re.IGNORECASE):
            trecho = texto[max(0, m.start() - 20): min(len(texto), m.end() + janela)]
            pares.extend(_DATA_RANGE_RE.findall(trecho))
            pares.extend(_DATA_INICIO_TERMINO_RE.findall(trecho))
        if pares:
            return pares
    return []


def _periodo_do_pdf(texto: str, nome_pesquisador: str = "") -> tuple[date, date] | None:
    """Extrai o período de um comprovante — retorna (menor início, maior
    fim).

    Prioridade: (1) "Contratos de trabalho" de Carteira de Trabalho
    Digital/Extrato de Vínculos (ver _pares_contratos_trabalho — formato
    de documento bem específico, então tem prioridade máxima quando bate;
    sem isso, datas de outras seções do mesmo documento, tipo "Férias",
    podiam ser confundidas com a duração do vínculo); (2) pares de data
    numa janela ancorada no nome do pesquisador, se `nome_pesquisador` for
    informado e aparecer no texto (ver _pares_ancorados_no_nome — evita
    misturar com o período de outra pessoa em documento que lista
    várias); (3) pares "tendo atuado de X até Y" (participação real da
    pessoa, ver _DATA_ATUACAO_RE); (4) padrão genérico "DD/MM/AAAA
    <conector> DD/MM/AAAA" (_DATA_RANGE_RE) em todo o texto. Documentos
    com mais de um intervalo do mesmo tipo (ex.: papéis/cargos diferentes
    no mesmo certificado) têm o período completo coberto pelo menor
    início e maior fim entre eles.

    Retorna None se não achar nenhum par de datas reconhecível.
    """
    pares = _pares_contratos_trabalho(texto)
    if not pares and nome_pesquisador:
        pares = _pares_ancorados_no_nome(texto, nome_pesquisador)
    if not pares:
        pares = _DATA_ATUACAO_RE.findall(texto) or _DATA_RANGE_RE.findall(texto)
    if not pares:
        return None
    datas_ini, datas_fim = [], []
    for d1, m1, a1, d2, m2, a2 in pares:
        try:
            datas_ini.append(date(int(a1), int(m1), int(d1)))
            datas_fim.append(date(int(a2), int(m2), int(d2)))
        except ValueError:
            continue
    if not datas_ini or not datas_fim:
        return None
    inicio, fim = min(datas_ini), max(datas_fim)
    # Data de fim no futuro não é uma duração real — é "previsão de
    # término"/conclusão de um projeto ainda em andamento. Não tem como
    # contar meses até uma data que ainda não chegou; usa hoje como fim.
    hoje = date.today()
    if fim > hoje:
        fim = hoje
    if fim < inicio:
        return None
    return inicio, fim


def _meses_entre(inicio: date, fim: date) -> int:
    """Meses entre duas datas, com precisão de dia — dias corridos ÷ 30,44
    (média real de dias por mês), arredondado pro inteiro mais próximo.

    Evita o viés de contar só ano/mês (ex.: 24/05 a 13/12 são só ~6,6
    meses corridos, mas "mês tocado, mesmo que 1 dia" contaria 8 — maio e
    dezembro inteiros mesmo participando só 8 e 13 dias deles,
    respectivamente). Ao mesmo tempo, continua dando o valor certo pra
    períodos de mês cheio (ex.: 01/09 a 31/08 do ano seguinte = 365 dias
    ≈ 12 meses, um ano de bolsa)."""
    dias = (fim - inicio).days
    return round(dias / 30.44)




# Seção do PDF comprovante por categoria de atuação, conforme a numeração do
# Anexo II (7 = docente ensino superior, 8 = docente ensino básico, 9 = não
# docente) — são três categorias com pontuação própria, cada uma com seu
# próprio bloco de PDFs na pasta do candidato, não uma seção única.
_SECAO_ATUACAO = {"ensino_superior": "7", "ensino_basico": "8", "nao_docencia": "9"}


def extrair_atuacao_profissional(root: ET.Element, pontos_atuacao: dict) -> list[dict]:
    items = []
    cnt = {"ensino_superior": 0, "ensino_basico": 0, "nao_docencia": 0}
    # Sigla/nome fantasia da instituição fica num elemento à parte
    # (INFORMACAO-ADICIONAL-INSTITUICAO), ligado por CODIGO-INSTITUICAO —
    # não vem junto do NOME-INSTITUICAO em ATUACAO-PROFISSIONAL. Importa
    # porque o comprovante (carteira de trabalho) costuma trazer a razão
    # social/CNPJ, que às vezes bate com a sigla e não com o nome completo
    # cadastrado no Lattes (ex.: Lattes = "Decathlon", sigla = "IGUASPORT" —
    # a razão social real da loja no Brasil).
    sigla_por_codigo = {
        info.get("CODIGO-INSTITUICAO", ""): info.get("SIGLA-INSTITUICAO", "")
        for info in root.findall(".//INFORMACAO-ADICIONAL-INSTITUICAO")
    }
    for ap in root.findall(".//ATUACAO-PROFISSIONAL"):
        instituicao = ap.get("NOME-INSTITUICAO", "")
        codigo      = ap.get("CODIGO-INSTITUICAO", "")
        sigla       = sigla_por_codigo.get(codigo, "")
        for vinculo in ap.findall("VINCULOS"):
            cargo      = vinculo.get("OUTRO-ENQUADRAMENTO-FUNCIONAL-INFORMADO", "")
            outro_vinc = vinculo.get("OUTRO-VINCULO-INFORMADO", "")
            mes_ini    = vinculo.get("MES-INICIO", "")
            ano_ini    = vinculo.get("ANO-INICIO", "")
            mes_fim    = vinculo.get("MES-FIM", "")
            ano_fim    = vinculo.get("ANO-FIM", "")
            if not ano_ini:
                continue
            descricao = cargo or outro_vinc or "—"
            categoria = classificar_atividade(instituicao, descricao)
            meses    = calcular_meses(mes_ini, ano_ini, mes_fim, ano_fim)
            pts_unit = pontos_atuacao.get(categoria, 0)

            periodo_ini = f"{mes_ini.zfill(2)}/{ano_ini}" if mes_ini else ano_ini
            periodo_fim = (f"{mes_fim.zfill(2)}/{ano_fim}" if mes_fim else ano_fim) if ano_fim else "atual"

            cnt[categoria] = cnt.get(categoria, 0) + 1
            secao = _SECAO_ATUACAO.get(categoria, "9")
            items.append({
                "seq":         f"{secao}.{cnt[categoria]}",
                "instituicao": instituicao,
                "instituicao_sigla":  sigla,
                "instituicao_codigo": codigo,
                "descricao":   descricao,
                "periodo":     f"{periodo_ini} – {periodo_fim}",
                "ano_inicio":  ano_ini,
                "ano_fim":     ano_fim,
                "categoria":   categoria,
                "meses":       meses,
                "pontos_unit": pts_unit,
                "pontos":      meses * pts_unit,
            })
    return items


_VIGENTE_KW = ("atual", "presente data", "momento", "vigente")

# Carteira de trabalho usa nomenclatura CBO abreviada/genérica ("VENDEDOR
# TECNICO JR"), enquanto o Lattes costuma trazer a forma completa e no
# gênero da pessoa ("Vendedora Júnior") — sem normalizar isso, a
# comparação de texto nunca bate mesmo quando é claramente o mesmo cargo.
_ABREV_CARGO = [
    (r"\bsr\b",    "senior"),
    (r"\bjr\b",    "junior"),
    (r"\bpl\b",    "pleno"),
    (r"\btec\b",   "tecnico"),
    (r"\btecn\b",  "tecnico"),
    (r"\bassist\b", "assistente"),
    (r"\bcoord\b", "coordenador"),
]


def _normalizar_cargo(texto: str) -> str:
    """Normaliza texto de cargo pra comparação: expande abreviações de
    nível comuns em CBO (Sr/Jr/Pl → sênior/júnior/pleno) e neutraliza
    gênero gramatical (tira o "a"/"o" final de palavras longas o
    bastante — "vendedora"/"vendedor" e "tecnica"/"tecnico" caem no
    mesmo radical). Aplicado nos dois lados da comparação."""
    t = normalizar(texto)
    for pat, rep in _ABREV_CARGO:
        t = re.sub(pat, rep, t)
    palavras = []
    for p in re.findall(r"\w+", t):
        if len(p) > 4 and p[-1] in "ao":
            p = p[:-1]
        palavras.append(p)
    return " ".join(palavras)


def similaridade_cargo(cargo: str, texto_pdf: str) -> float:
    """Como similaridade(), mas usa _normalizar_cargo() nos dois lados em
    vez de normalizar() puro — ver _normalizar_cargo."""
    stop = {"a", "o", "e", "de", "do", "da", "em", "no", "na",
            "the", "of", "in", "and", "for", "to", "with"}
    palavras = [p for p in re.findall(r"\w+", _normalizar_cargo(cargo)) if len(p) > 3 and p not in stop]
    if not palavras:
        return 0.0
    tokens_b = set(re.findall(r"\w+", _normalizar_cargo(texto_pdf)))
    return sum(1 for p in palavras if p in tokens_b) / len(palavras)


# Mais permissivo que LIMIAR_PERIODICO (usado pra título de periódico/
# evento, textos "limpos"): comprovante de vínculo empregatício é bem mais
# ruidoso — CBO abreviado, nome de cargo/setor com sinônimos, histórico de
# promoções que o pesquisador pode descrever diferente do documento oficial.
LIMIAR_CARGO_ATUACAO = 0.30


def verificar_atuacao(item: dict, texto_pdf: str, nome_pesquisador: str) -> dict:
    texto_norm = normalizar(texto_pdf)

    partes    = normalizar(nome_pesquisador).split()
    sobrenome = partes[-1] if partes else ""
    nome_ok   = bool(sobrenome) and sobrenome in texto_norm

    sim_inst = similaridade_instituicao(item["instituicao"], texto_pdf)
    # A carteira de trabalho costuma trazer a razão social/CNPJ do
    # empregador, que às vezes não tem nada a ver com o nome que o
    # pesquisador cadastrou no Lattes (nome fantasia) — mas bate com a
    # sigla/nome fantasia cadastrado à parte (ver extrair_atuacao_profissional).
    # Sigla curta demais (≤3 chars, tipo "FS") é ignorada — risco alto de
    # coincidência.
    sigla    = item.get("instituicao_sigla", "")
    sigla_ok = len(sigla) > 3 and normalizar(sigla) in texto_norm
    inst_ok  = sim_inst >= LIMIAR_INSTITUICAO or sigla_ok

    cargo = item.get("descricao", "")
    if cargo and cargo != "—":
        sim_cargo = similaridade_cargo(cargo, texto_pdf)
        cargo_ok  = sim_cargo >= LIMIAR_CARGO_ATUACAO
    else:
        sim_cargo, cargo_ok = None, True  # XML sem cargo informado — nada a conferir

    ano_ini = item.get("ano_inicio", "")
    ini_ok  = bool(ano_ini) and ano_ini in texto_pdf

    ano_fim = item.get("ano_fim", "")
    if ano_fim:
        fim_ok = ano_fim in texto_pdf
    else:
        fim_ok = any(kw in texto_norm for kw in _VIGENTE_KW)
    per_ok = ini_ok and fim_ok

    aprovado    = nome_ok and inst_ok and cargo_ok and per_ok
    cargo_label = f"{sim_cargo:.0%}" if sim_cargo is not None else "—"
    inst_label  = f"{sim_inst:.0%}" + (" (sigla ✓)" if sigla_ok else "")
    detalhes = (f"Nome: {'✓' if nome_ok else '✗'} | "
                f"Inst.: {inst_label} | "
                f"Cargo: {cargo_label} | "
                f"Período: {'✓' if per_ok else '✗'}")
    # Score pra desempate quando mais de um vínculo do Lattes passa no
    # limiar pro mesmo PDF (ver _atribuicao_otima em verificar_curriculo).
    # Sigla batendo é um sinal forte — vale o mesmo que 100% de nome.
    score = max(sim_inst, 1.0 if sigla_ok else 0.0)
    return {"aprovado": aprovado, "detalhes": detalhes, "score": score}


def _agrupar_atuacao_por_instituicao(atuacoes_xml: list[dict]) -> list[dict]:
    """Agrupa os vínculos de atuação por instituição (+ categoria, pro caso
    raro de a mesma instituição ter vínculos classificados diferente —
    ex.: começou não-docente, virou docente). Uma carteira de trabalho
    prova o vínculo com o empregador como um todo, não uma fase/promoção
    específica — então o casamento com PDF é feito por instituição, não
    por vínculo individual (ver verificar_curriculo)."""
    grupos: dict[tuple, dict] = {}
    ordem = []
    for item in atuacoes_xml:
        chave = (item.get("instituicao_codigo") or item["instituicao"], item["categoria"])
        if chave not in grupos:
            grupos[chave] = {
                "instituicao": item["instituicao"],
                "sigla":       item.get("instituicao_sigla", ""),
                "categoria":   item["categoria"],
                "itens":       [],
            }
            ordem.append(chave)
        grupos[chave]["itens"].append(item)
    return [grupos[k] for k in ordem]


def _verificar_atuacao_instituicao(grupo: dict, texto_pdf: str, nome_pesquisador: str) -> dict:
    """Como verificar_atuacao(), mas só confirma nome do pesquisador +
    instituição/sigla — sem checar cargo nem período de um vínculo
    específico. Aqui a ideia é provar o vínculo com o empregador como um
    todo (o PDF pode cobrir várias promoções/cargos de uma vez, cada uma
    com seu próprio cargo — checar cargo de UM vínculo contra o PDF
    inteiro não faz sentido nesse nível); o período/meses creditados vêm
    de _periodo_do_pdf, não daqui."""
    texto_norm = normalizar(texto_pdf)
    partes    = normalizar(nome_pesquisador).split()
    sobrenome = partes[-1] if partes else ""
    nome_ok   = bool(sobrenome) and sobrenome in texto_norm

    sim_inst = similaridade_instituicao(grupo["instituicao"], texto_pdf)
    sigla    = grupo.get("sigla", "")
    sigla_ok = len(sigla) > 3 and normalizar(sigla) in texto_norm
    inst_ok  = sim_inst >= LIMIAR_INSTITUICAO or sigla_ok

    aprovado   = nome_ok and inst_ok
    inst_label = f"{sim_inst:.0%}" + (" (sigla ✓)" if sigla_ok else "")
    detalhes   = f"Nome: {'✓' if nome_ok else '✗'} | Inst.: {inst_label}"
    score      = max(sim_inst, 1.0 if sigla_ok else 0.0)
    return {"aprovado": aprovado, "detalhes": detalhes, "score": score}


# ── Verificadores por tipo ────────────────────────────────────────────────────

LIMIAR_TITULO       = 0.60
LIMIAR_PERIODICO    = 0.50
LIMIAR_INSTITUICAO  = 0.75  # mais rígido: nomes de instituição são curtos,
                            # cada palavra pesa muito na conta de similaridade


def _doi_no_pdf(doi: str, texto_pdf: str) -> bool:
    if not doi or not texto_pdf:
        return False
    doi_norm = re.sub(r"https?://(dx\.)?doi\.org/", "", doi.strip().lower())
    return doi_norm in texto_pdf.lower()


# ── Checadores atômicos ───────────────────────────────────────────────────────

def _checar_titulo(item, texto, _nome=""):
    sim = similaridade(item.get("titulo", ""), texto)
    return sim >= LIMIAR_TITULO, f"Título: {sim:.0%}"

def _checar_doi(item, texto, _nome=""):
    ok = _doi_no_pdf(item.get("doi", ""), texto)
    return ok, f"DOI: {'✓' if ok else '✗'}"

def _checar_issn(item, texto, _nome=""):
    issn = item.get("issn", "").replace("-", "")
    ok = bool(issn) and issn in texto.replace("-", "")
    return ok, f"ISSN: {'✓' if ok else '✗'}"

def _checar_autor(item, texto, nome=""):
    partes = normalizar(nome).split()
    sob = partes[-1] if partes else ""
    ok = bool(sob) and sob in normalizar(texto)
    return ok, f"Nome: {'✓' if ok else '✗'}"

def _checar_periodico(item, texto, _nome=""):
    sim = similaridade(item.get("periodico", ""), texto)
    return sim >= LIMIAR_PERIODICO, f"Periódico: {sim:.0%}"

def _checar_evento(item, texto, _nome=""):
    sim = similaridade(item.get("evento", ""), texto)
    return sim >= LIMIAR_PERIODICO, f"Evento: {sim:.0%}"

def _checar_isbn(item, texto, _nome=""):
    isbn = item.get("isbn", "").replace("-", "").replace(" ", "")
    ok = bool(isbn) and isbn in texto.replace("-", "").replace(" ", "")
    return ok, f"ISBN: {'✓' if ok else '✗'}"

def _checar_orientando(item, texto, _nome=""):
    parts = item.get("orientando", "").split()
    sob = parts[-1] if parts else ""
    ok = bool(sob) and normalizar(sob) in normalizar(texto)
    return ok, f"Orientando: {'✓' if ok else '✗'}"

def _checar_candidato(item, texto, _nome=""):
    parts = item.get("candidato", "").split()
    sob = parts[-1] if parts else ""
    ok = bool(sob) and normalizar(sob) in normalizar(texto)
    return ok, f"Candidato: {'✓' if ok else '✗'}"

def _checar_instituicao(item, texto, _nome=""):
    sim = similaridade_instituicao(item.get("instituicao", ""), texto)
    return sim >= LIMIAR_INSTITUICAO, f"Instituição: {sim:.0%}"

def _checar_periodo(item, texto, _nome=""):
    ano = item.get("ano", item.get("ano_inicio", ""))
    ok = bool(ano) and ano in texto
    return ok, f"Período: {'✓' if ok else '✗'}"

def _checar_entidade(item, texto, _nome=""):
    sim = similaridade(item.get("entidade", ""), texto)
    return sim >= LIMIAR_PERIODICO, f"Entidade: {sim:.0%}"

def _checar_editora(item, texto, _nome=""):
    sim = similaridade(item.get("editora", ""), texto)
    return sim >= LIMIAR_PERIODICO, f"Editora: {sim:.0%}"


_CHECADORES = {
    "titulo":      _checar_titulo,
    "doi":         _checar_doi,
    "issn":        _checar_issn,
    "autor":       _checar_autor,
    "periodico":   _checar_periodico,
    "evento":      _checar_evento,
    "isbn":        _checar_isbn,
    "orientando":  _checar_orientando,
    "candidato":   _checar_candidato,
    "instituicao": _checar_instituicao,
    "periodo":     _checar_periodo,
    "entidade":    _checar_entidade,
    "editora":     _checar_editora,

}

_VOCAB_CAMPOS = {
    "titulo": "titulo", "doi": "doi", "issn": "issn",
    "autor": "autor", "periodico": "periodico", "evento": "evento",
    "isbn": "isbn", "orientando": "orientando", "candidato": "candidato",
    "instituicao": "instituicao", "instituição": "instituicao",
    "periodo": "periodo", "período": "periodo",
    "entidade": "entidade", "editora": "editora",
    "apenas_pdf": "apenas_pdf",
}


def _parse_campos_config(texto_col_f) -> list[str]:
    """Converte o valor da coluna F do Excel em lista de campos a verificar."""
    if not texto_col_f:
        return ["titulo"]
    campos = []
    for token in re.split(r"[+,;&\n]", normalizar(str(texto_col_f))):
        token = token.strip()
        if token in _VOCAB_CAMPOS:
            campos.append(_VOCAB_CAMPOS[token])
    return campos or ["titulo"]


def verificar_por_config(item: dict, texto: str,
                         campos: list[str], nome_pesquisador: str = "") -> dict:
    """Verificador genérico: avalia campos definidos no Excel (lógica OR)."""
    if "apenas_pdf" in campos:
        return {"aprovado": True, "detalhes": "Apenas presença do PDF verificada."}
    aprovado = False
    dets: list[str] = []
    for campo in campos:
        fn = _CHECADORES.get(campo)
        if fn is None:
            continue
        ok, det = fn(item, texto, nome_pesquisador)
        dets.append(det)
        if ok:
            aprovado = True
    return {"aprovado": aprovado, "detalhes": " | ".join(dets) or "Nenhum campo configurado."}


# ── Pontuação por item de produção ────────────────────────────────────────────

_CIDADES_BR = {
    "sao paulo", "rio de janeiro", "porto alegre", "belo horizonte",
    "brasilia", "brasília", "campinas", "curitiba", "florianopolis",
    "florianópolis", "recife", "salvador", "fortaleza", "goiania",
    "goiânia", "manaus", "natal", "vitoria", "vitória", "belem", "belém",
    "maceio", "maceió", "joao pessoa", "joão pessoa", "teresina",
    "campo grande", "cuiaba", "cuiabá", "macapa", "macapá",
    "porto velho", "boa vista", "palmas", "aracaju", "sao luis",
    "são luís", "rio branco",
}


def _classif_evento(classif: str) -> str:
    c = classif.upper()
    if "INTERNAC" in c:
        return "Internacional"
    if "NACION" in c:
        return "Nacional"
    return "Regional"


def _cidade_nacional(cidade: str) -> bool:
    return normalizar(cidade) in _CIDADES_BR


def calcular_pontos_item(item: dict, tipo: str,
                          criterios: dict, qualis: dict) -> tuple[float, str]:
    """Retorna (pontos, subcategoria_determinada)."""
    c = criterios

    if tipo == "artigo":
        issn    = item.get("issn", "").replace("-", "").strip()
        estrato = qualis.get(issn) if issn else None
        if estrato is None:
            # ISSN do Lattes não bateu — acontece na prática quando o
            # pesquisador cadastrou o ISSN de uma versão antiga/impressa do
            # periódico e o Qualis usa outro (ex.: online). Antes de desistir
            # e cair em "Sem Qualis", tenta pelo nome exato do periódico e,
            # se ainda assim não achar, por nome aproximado (ex.: "Revista
            # Cena" no Lattes x "Cena. UFRGS" no Qualis — mesmo periódico,
            # nome cadastrado diferente).
            periodico = item.get("periodico", "")
            estrato = (qualis.get(_norm_periodico(periodico))
                       or _qualis_por_nome_aproximado(periodico, qualis.get("__titulos__", []))
                       or "Sem Qualis")
        tabela  = c.get("artigo", {})
        autoria = item.get("autoria", "")
        # Se a planilha distinguir Primeiro/Demais (chave "A1-Primeiro" —
        # ver _resolver_sub), usa a taxa certa pra esse autor específico;
        # cai pro nível "flat" (só "A1") se a planilha não fizer essa
        # distinção, ou se a autoria não foi identificada no Lattes.
        if autoria:
            chave = f"{estrato}-{autoria}"
            if chave in tabela:
                return tabela[chave], chave
        return tabela.get(estrato, 0), estrato

    if tipo in ("trabalho_completo", "resumo"):
        tabela = c.get(tipo, {})
        if "unidade" in tabela:
            return tabela["unidade"], "Por unidade"
        sub = _classif_evento(item.get("classificacao", ""))
        return tabela.get(sub, 0), sub

    if tipo in ("livro", "livro_publicado"):
        tipo_liv = item.get("tipo", "")
        sub_tipo = "livro_organizado" if "ORGANIZADO" in tipo_liv.upper() or "EDICAO" in tipo_liv.upper() \
                   else "livro_publicado"
        tabela = c.get(sub_tipo, {})
        if "unidade" in tabela:
            return tabela["unidade"], "Por unidade"
        sub_loc = "Nacional" if _cidade_nacional(item.get("cidade_editora", "")) else "Internacional"
        return tabela.get(sub_loc, 0), sub_loc

    if tipo == "capitulo":
        tabela  = c.get("capitulo", {})
        autoria = item.get("autoria", "")
        # Se a planilha distinguir Autor/Coautor (chave "Autor"/"Coautor" —
        # ver _resolver_sub), usa a taxa certa; cai pro "unidade"/
        # Nacional-Internacional se a planilha não fizer essa distinção,
        # ou se a autoria não foi identificada no Lattes.
        if autoria and autoria in tabela:
            return tabela[autoria], autoria
        if "unidade" in tabela:
            return tabela["unidade"], "Por unidade"
        sub = "Nacional" if _cidade_nacional(item.get("cidade_editora", "")) else "Internacional"
        return tabela.get(sub, 0), sub

    if tipo == "orientacao":
        nat  = item.get("natureza", "")
        role = item.get("tipo_orientacao", "ORIENTADOR_PRINCIPAL")
        sub  = "principal" if "PRINCIPAL" in role.upper() else "coorientador"
        if nat == "Mestrado":
            return c.get("orientacao_mestrado", {}).get(sub, 0), f"Mestrado/{sub}"
        nat_map = {
            "TCC Graduação": "orientacao_tcc",
            "Iniciação Científica": "orientacao_ic",
            "Especialização": "orientacao_esp",
        }
        chave  = nat_map.get(nat, "orientacao_tcc")
        tabela = c.get(chave, {})
        return tabela.get(sub, tabela.get("principal", 0)), f"{nat}/{sub}"

    if tipo == "banca":
        # Nem todo edital distingue nível de banca (o Anexo II atual só tem
        # uma linha, "participação em bancas de TCC", sem separar
        # graduação/especialização/mestrado/doutorado) — cai pra
        # "Graduacao" quando não existe taxa específica pro nível, mesmo
        # padrão de fallback já usado em projeto_pesquisa/extensao.
        sub    = _sub_banca(item.get("tipo", ""))
        tabela = c.get("banca", {})
        return tabela.get(sub, tabela.get("Graduacao", 0)), sub

    if tipo == "banca_especializacao_manual":
        return c.get("banca", {}).get("Especializacao", 0), "Especializacao"

    if tipo in ("projeto_pesquisa", "projeto_extensao"):
        # Nem todo edital distingue coordenador de integrante nessas duas
        # categorias (o Anexo II atual não distingue — uma taxa única por
        # mês) — quando a planilha só tem "integrante" cadastrado, cai pra
        # ela em vez de zerar coordenador por falta de chave específica
        # (mesmo padrão de fallback já usado em orientação, logo abaixo).
        papel  = item.get("papel", "integrante")
        tabela = c.get(tipo, {})
        taxa   = tabela.get(papel, tabela.get("integrante", 0))
        # Pontuação é por MÊS (ver extrair_projetos) — sem multiplicar por
        # "meses", um projeto de 2 anos valeria o mesmo que um de 2 meses.
        meses  = item.get("meses", 0)
        return taxa * meses, f"{papel.capitalize()} ({meses}m)"

    if tipo in ("organizacao_evento", "ic_manual"):
        return c.get(tipo, {}).get("unidade", 0), "Por unidade"

    if tipo == "prod_artistica_int":
        return c.get("prod_artistica_int", {}).get("unidade", 0), "Internacional"
    if tipo == "prod_artistica_nac":
        return c.get("prod_artistica_nac", {}).get("unidade", 0), "Nacional"
    if tipo == "prod_artistica_reg":
        return c.get("prod_artistica_reg", {}).get("unidade", 0), "Regional/Local"

    return 0, "—"


# ── Localizar PDF comprovante ─────────────────────────────────────────────────

PASTA_POR_CRITERIO = {
    "1": "1_artigos", "2": "2_trabalhos_completos", "3": "3_resumos",
    "4": "4_livros",  "5": "5_capitulos",           "6": "6_orientacoes",
    "7": "7_bancas",                                   "8": "8_atuacao",
    "9": "9_projetos_pesquisa",  "10": "10_projetos_extensao",
    "13": "13_organizacao_evento",
    "14": "14_iniciacao_cientifica",
    "15": "15_banca_especializacao",
    "16": "16_prod_artistica_internacional",
    "17": "17_prod_artistica_nacional",
    "18": "18_prod_artistica_regional",
}


def localizar_pdf(pasta_candidato: Path, seq: str) -> Path | None:
    # Estrutura plana (nova): PDFs direto na pasta do candidato
    for ext in (".pdf", ".PDF"):
        p = pasta_candidato / f"{seq}{ext}"
        if p.exists():
            return p
    # Estrutura com subpastas (legado)
    criterio = seq.split(".")[0]
    subpasta  = pasta_candidato / "comprovantes" / PASTA_POR_CRITERIO.get(criterio, "")
    for ext in (".pdf", ".PDF"):
        p = subpasta / f"{seq}{ext}"
        if p.exists():
            return p
    return None


def _pool_pdfs_secao(pasta: Path, secao: str) -> list[tuple[Path, str]]:
    """Coleta e pré-extrai texto de todos os PDFs de uma seção."""
    pool = []
    candidatos = sorted(
        list(pasta.glob(f"{secao}.*.pdf")) +
        list(pasta.glob(f"{secao}.*.PDF"))
    )
    for pdf in candidatos:
        if pdf.stem.split(".")[0] == str(secao):
            pool.append((pdf, extrair_texto_pdf(pdf)))
    return pool


# Campos que identificam univocamente a qual item um PDF pertence.
# "autor" é excluído: o sobrenome do pesquisador aparece em todos os seus PDFs
# e causaria falsos positivos ao casar PDFs com itens errados.
_CAMPOS_ID = frozenset({
    "titulo", "doi", "isbn", "issn", "evento",
    "orientando", "candidato", "entidade", "editora",
})


def _numero_pdf(pdf: Path) -> tuple[int, ...]:
    """Converte '1.3.pdf' em (1, 3) para ordenação natural."""
    try:
        return tuple(int(p) for p in pdf.stem.split("."))
    except ValueError:
        return (9999,)


def _score_match(item: dict, texto: str, campos_id: list[str], nome: str) -> float:
    """Retorna score de correspondência entre um item Lattes e o texto de um PDF.

    DOI/ISBN exatos → 2.0 (match definitivo).
    Similaridade de título → valor em [0, 1].
    Outros campos (issn, evento…) → 0.5 se presentes (sinal fraco).
    """
    melhor = 0.0
    for campo in campos_id:
        if campo == "titulo":
            s = similaridade(item.get("titulo", ""), texto)
            melhor = max(melhor, s)
        elif campo == "doi":
            if _doi_no_pdf(item.get("doi", ""), texto):
                return 2.0
        elif campo == "isbn":
            isbn = item.get("isbn", "").replace("-", "")
            if isbn and isbn in texto.replace("-", ""):
                return 2.0
        else:
            fn = _CHECADORES.get(campo)
            if fn:
                ok, _ = fn(item, texto, nome)
                if ok:
                    melhor = max(melhor, 0.5)
    return melhor


def _atribuicao_otima(matriz_scores: list[list[float]], limiar: float) -> dict[int, tuple[int | None, float]]:
    """Casa PDFs (linhas) com itens do Lattes (colunas) maximizando a SOMA
    total dos scores — não "cada PDF pega guloso o seu melhor item
    individual". A diferença importa quando dois itens têm títulos quase
    idênticos (ex.: mesma atividade em edições/anos diferentes, tipo
    "Edição 2022" vs. "Edição 2023"): um casamento guloso pode deixar o
    primeiro PDF roubar o item de um segundo PDF que combinaria melhor com
    ele, derrubando o total; a atribuição ótima não permite isso — se dois
    PDFs trocados batem melhor entre si, é essa troca que vence.

    Usa o algoritmo húngaro (scipy) quando disponível; sem scipy, cai para
    guloso por score decrescente (globalmente pior em casos raros de
    empate, mas ainda assim resolve o caso comum acima, já que o par de
    maior score é sempre atribuído primeiro).

    Retorna {índice_da_linha: (índice_da_coluna ou None, score)} — só
    inclui pares com score >= limiar; abaixo disso conta como "sem match".
    """
    n_pdf  = len(matriz_scores)
    n_item = len(matriz_scores[0]) if matriz_scores else 0
    resultado: dict[int, tuple[int | None, float]] = {i: (None, 0.0) for i in range(n_pdf)}
    if n_pdf == 0 or n_item == 0:
        return resultado

    if _HAS_SCIPY:
        m = np.array(matriz_scores, dtype=float)
        linhas, colunas = linear_sum_assignment(m, maximize=True)
        for r, c in zip(linhas, colunas):
            if m[r, c] >= limiar:
                resultado[r] = (int(c), float(m[r, c]))
        return resultado

    pares = sorted(
        ((matriz_scores[i][j], i, j) for i in range(n_pdf) for j in range(n_item)
         if matriz_scores[i][j] >= limiar),
        key=lambda t: -t[0],
    )
    usados_item: set[int] = set()
    for score, i, j in pares:
        if resultado[i][0] is not None or j in usados_item:
            continue
        usados_item.add(j)
        resultado[i] = (j, score)
    return resultado


def _processar_secao_por_pdf(
    items: list[dict],
    pool: list[tuple[Path, str]],
    campo_compl: str,
    tipo_pontos: str,
    campos_por_sub: dict,
    nome: str,
    criterios: dict,
    qualis: dict,
) -> list[dict]:
    """Processa uma seção partindo dos PDFs disponíveis (não dos itens do Lattes).

    - Casamento PDF↔item é feito por atribuição ÓTIMA (ver _atribuicao_otima):
      maximiza a soma dos scores de todos os PDFs da seção juntos, não o
      melhor score de cada PDF isoladamente — evita que um PDF "roube" o
      item certo de outro PDF quando dois itens têm títulos quase idênticos
      (ex.: mesma atividade em edições/anos diferentes).
    - PDFs sem casamento → REPROVADO.
    - Itens sem PDF → omitidos do relatório.
    - Ordem: numeração dos PDFs (1.1.pdf, 1.2.pdf, …).
    - `campos_por_sub` mapeia subcategoria → campos a verificar (ex.: banca
      "Graduacao" pede "candidato", "Mestrado" pede "orientando"); a
      subcategoria de cada item só é conhecida após o casamento com o PDF,
      então o campo de identificação usado para casar (campos_id) é a união
      de todas as subcategorias do tipo.
    """
    # Se a planilha de critérios não configurou "O que verificar no PDF"
    # pra esse tipo (coluna em branco — acontece, nem toda linha do
    # criterios.xlsx preenche isso), cai pros campos padrão do Python em
    # vez de ir direto pro fallback genérico ["titulo"]. Faz diferença de
    # verdade: capítulo/livro têm ISBN como identificador muito mais
    # confiável que título quando o PDF anexado é o livro inteiro (às
    # vezes centenas de páginas) — o ISBN geralmente aparece já nas
    # primeiras páginas (capa/ficha catalográfica), então "titulo" sozinho
    # falha (só as 3 primeiras páginas são lidas, e o título do capítulo
    # em si pode estar bem mais adiante no livro).
    campos_por_sub_efetivo = campos_por_sub or _campos_config_padrao().get(tipo_pontos, {})
    campos_todos = {c for lst in campos_por_sub_efetivo.values() for c in lst} if campos_por_sub_efetivo else set()
    campos_id    = [c for c in campos_todos if c in _CAMPOS_ID] or ["titulo"]
    resultados   = []

    pool_validos = []
    for pdf, texto in pool:
        if not texto.strip() or texto.startswith("__ERRO_PDF__"):
            resultados.append({
                "seq":          pdf.stem,
                "titulo":       "—",
                "complemento":  "",
                "ano":          "",
                "status":       "ERRO PDF" if texto.startswith("__ERRO_PDF__") else "SEM TEXTO",
                "detalhes":     texto or "PDF sem texto extraível.",
                "pdf_nome":     pdf.name,
                "pontos":       0,
                "subcategoria": "—",
                "tipo_pontos":  tipo_pontos,
            })
        else:
            pool_validos.append((pdf, texto))

    matriz      = [[_score_match(item, texto, campos_id, nome) for item in items]
                   for _, texto in pool_validos]

    # PDF "quase bateu" (score logo abaixo do limiar) pode ser porque só as
    # 3 primeiras páginas foram lidas, mas o arquivo anexado é bem maior
    # que o comprovante em si — candidato às vezes sobe o LIVRO inteiro em
    # vez de só o capítulo, ou o CADERNO DE RESUMOS do congresso inteiro em
    # vez de só o resumo dela; o título/ISBN relevante pode estar bem mais
    # adiante no arquivo. Tenta de novo com mais páginas só nesse caso
    # (custo extra só quando parece valer a pena — não vale reler todo PDF
    # da seção, a maioria já bate ou não bate claramente nas 3 primeiras).
    JANELA_QUASE_BATEU = 0.15
    PAGINAS_RETRY      = 50
    textos_completos: dict[int, str] = {}
    for i, (pdf, texto) in enumerate(pool_validos):
        for j, item in enumerate(items):
            score = matriz[i][j]
            if LIMIAR_TITULO - JANELA_QUASE_BATEU <= score < LIMIAR_TITULO:
                if i not in textos_completos:
                    textos_completos[i] = extrair_texto_pdf(pdf, max_paginas=PAGINAS_RETRY)
                texto_completo = textos_completos[i]
                if texto_completo != texto:
                    novo_score = _score_match(item, texto_completo, campos_id, nome)
                    if novo_score > score:
                        matriz[i][j] = novo_score
    # Troca o texto "oficial" desses PDFs pela versão com mais páginas —
    # daqui pra frente (verificação de campos, extração de período etc.)
    # todo mundo enxerga o texto mais completo, não só o score.
    for i, texto_completo in textos_completos.items():
        pool_validos[i] = (pool_validos[i][0], texto_completo)

    atribuicao  = _atribuicao_otima(matriz, LIMIAR_TITULO)

    # Projeto de pesquisa/extensão: às vezes o Lattes só tem 1 item pra uma
    # atividade plurianual (ano_fim > ano_inicio) que gerou mais de um
    # comprovante — um PDF por edição/período (ex.: "IV Programa..." e
    # "VIII Programa..." do mesmo projeto, mesmo item do Lattes). Depois da
    # atribuição 1-pra-1 normal, qualquer PDF que sobrou mas também bate MUITO
    # bem (>= LIMIAR_EDICAO_EXTRA, mais rígido que o limiar normal de
    # aprovação) com um item JÁ atribuído a outro PDF é tratado como edição
    # extra do MESMO item, em vez de reprovado — os períodos de todos os
    # PDFs daquele item entram juntos no cálculo de meses (_periodo_do_pdf
    # sobre os textos concatenados). Limiar mais alto que o normal (0.60) de
    # propósito: aqui o risco de falso positivo é maior — dois projetos de
    # temas parecidos (ex.: "treinamento físico") podem passar de 60% por
    # vocabulário genérico do domínio sem serem o mesmo projeto; herdar
    # pontos do item errado é pior que só reprovar por segurança.
    LIMIAR_EDICAO_EXTRA = 0.85
    extras_por_item: dict[int, list[int]] = {}
    if tipo_pontos in ("projeto_pesquisa", "projeto_extensao"):
        usados = {item_idx for item_idx, _ in atribuicao.values() if item_idx is not None}
        for i in range(len(pool_validos)):
            if atribuicao[i][0] is not None or not usados:
                continue
            melhor_j, melhor_score = None, LIMIAR_EDICAO_EXTRA
            for j in usados:
                if matriz[i][j] >= melhor_score:
                    melhor_j, melhor_score = j, matriz[i][j]
            if melhor_j is not None:
                extras_por_item.setdefault(melhor_j, []).append(i)
    extras_pdf_idx = {i for lst in extras_por_item.values() for i in lst}

    # Banca: um único PDF pode reunir vários pareceres/bancas diferentes
    # num só documento (ex.: "Declaração de Participação" listando N
    # estudantes examinados) — cada um é um item DISTINTO do Lattes, e cada
    # um vale seus próprios pontos (diferente do caso de projeto acima, que
    # é o MESMO item em edições diferentes). Item que não ganhou a
    # atribuição 1-pra-1 principal, mas ainda bate muito bem com um PDF já
    # usado por OUTRO item, é aprovado também, na mesma linha do PDF.
    extras_item_por_pdf: dict[int, list[int]] = {}
    if tipo_pontos == "banca":
        pdfs_usados  = {i for i, (item_idx, _) in atribuicao.items() if item_idx is not None}
        itens_usados = {item_idx for item_idx, _ in atribuicao.values() if item_idx is not None}
        for j in range(len(items)):
            if j in itens_usados:
                continue
            melhor_i, melhor_score = None, LIMIAR_EDICAO_EXTRA
            for i in pdfs_usados:
                if matriz[i][j] >= melhor_score:
                    melhor_i, melhor_score = i, matriz[i][j]
            if melhor_i is not None:
                extras_item_por_pdf.setdefault(melhor_i, []).append(j)

    for i, (pdf, texto) in enumerate(pool_validos):
        item_idx, score = atribuicao[i]

        if item_idx is None and i in extras_pdf_idx:
            continue  # vira edição extra do PDF principal do item, abaixo

        if item_idx is not None:
            item_match = items[item_idx]
            extras_i   = extras_por_item.get(item_idx, [])
            if tipo_pontos in ("projeto_pesquisa", "projeto_extensao"):
                # Prioriza o período exato do(s) PDF(s) (o comprovante
                # costuma ter data completa) sobre a estimativa por ano
                # cheio do Lattes — só mantém a estimativa se nenhum PDF
                # do item tiver data reconhecível. Junta o texto de
                # eventuais edições extras pra achar o período completo.
                textos_item = [texto] + [pool_validos[k][1] for k in extras_i]
                periodo = _periodo_do_pdf("\n".join(textos_item), nome)
                if periodo is not None:
                    item_match["meses"] = _meses_entre(*periodo)
            pontos_item, sub_item = calcular_pontos_item(
                item_match, tipo_pontos, criterios, qualis)
            campos   = _resolver_campos(campos_por_sub_efetivo, sub_item)
            res_full = verificar_por_config(item_match, texto, campos, nome)
            # Artigo: mostra "Primeiro"/"Demais" junto do ano (coluna "Ano /
            # Autoria") — não pontua diferente hoje (criterios.xlsx não
            # distingue), mas ajuda a revisão manual a ver rápido.
            ano_exib = item_match.get("ano", "")
            autoria  = item_match.get("autoria", "")
            if autoria:
                ano_exib = f"{ano_exib} ({autoria})"
            # extrair_livros trata livro publicado/organizado como um único
            # tipo_pontos ("livro_publicado") — a distinção real só existe no
            # atributo TIPO do item, e calcular_pontos_item já usa essa mesma
            # regra pra escolher a tabela de pontos. Refaz aqui só pra
            # classificar certo na aba Resumo (seções 3 e 4 do Anexo II).
            tipo_pontos_real = tipo_pontos
            if tipo_pontos == "livro_publicado":
                tipo_liv = item_match.get("tipo", "").upper()
                if "ORGANIZADO" in tipo_liv or "EDICAO" in tipo_liv:
                    tipo_pontos_real = "livro_organizado"
            resultados.append({
                "seq":          pdf.stem,
                "titulo":       item_match["titulo"],
                "complemento":  item_match.get(campo_compl, ""),
                "ano":          ano_exib,
                "status":       "APROVADO",
                "detalhes":     res_full["detalhes"],
                "pdf_nome":     pdf.name,
                "pontos":       pontos_item,
                "subcategoria": sub_item,
                "tipo_pontos":  tipo_pontos_real,
                "meses":        item_match.get("meses"),
            })

            # Edições extras do mesmo item: aprovadas também (comprovante
            # genuíno), mas sem pontuar de novo — o período delas já entrou
            # no total calculado acima, na linha do PDF principal.
            for k in extras_i:
                pdf_extra, _ = pool_validos[k]
                resultados.append({
                    "seq":          pdf_extra.stem,
                    "titulo":       item_match["titulo"],
                    "complemento":  item_match.get(campo_compl, ""),
                    "ano":          ano_exib,
                    "status":       "APROVADO",
                    "detalhes":     (f"Edição adicional do mesmo item do Lattes — período já "
                                     f"somado à linha \"{pdf.stem}\" ({pdf.name})."),
                    "pdf_nome":     pdf_extra.name,
                    "pontos":       0,
                    "subcategoria": sub_item,
                    "tipo_pontos":  tipo_pontos_real,
                    # 0, não item_match["meses"] — os meses já foram
                    # contados na linha principal (pdf.stem, acima); somar
                    # de novo aqui duplicaria o total na aba Resumo, que
                    # soma "meses" de toda linha APROVADO da seção.
                    "meses":        0,
                })

            # Itens adicionais achados NO MESMO PDF (banca): cada um é uma
            # banca distinta, pontua cheio — sem relação com o mecanismo de
            # "edição extra" acima (que é o MESMO item repetido em PDFs
            # diferentes; aqui é o oposto, vários itens diferentes num só PDF).
            for j in extras_item_por_pdf.get(i, []):
                item_extra = items[j]
                pontos_extra, sub_extra = calcular_pontos_item(
                    item_extra, tipo_pontos, criterios, qualis)
                campos_extra = _resolver_campos(campos_por_sub_efetivo, sub_extra)
                res_extra    = verificar_por_config(item_extra, texto, campos_extra, nome)
                resultados.append({
                    "seq":          pdf.stem,
                    "titulo":       item_extra["titulo"],
                    "complemento":  item_extra.get(campo_compl, ""),
                    "ano":          item_extra.get("ano", ""),
                    "status":       "APROVADO",
                    "detalhes":     res_extra["detalhes"] + " (item adicional achado no mesmo PDF)",
                    "pdf_nome":     pdf.name,
                    "pontos":       pontos_extra,
                    "subcategoria": sub_extra,
                    "tipo_pontos":  tipo_pontos,
                    "meses":        item_extra.get("meses"),
                })
        else:
            # Diferencia "nada parecido o suficiente" de "o item mais
            # parecido foi para outro PDF que combinava ainda melhor com
            # ele" — a mensagem antiga ("abaixo do limiar") era enganosa
            # nesse segundo caso, porque o score podia estar bem acima do
            # limiar e mesmo assim não sobrar pro PDF que perdeu a disputa.
            melhor_score_linha = max(matriz[i], default=0.0)
            if melhor_score_linha >= LIMIAR_TITULO:
                j_melhor = matriz[i].index(melhor_score_linha)
                detalhes = (f"Melhor correspondência encontrada ('{items[j_melhor]['titulo'][:60]}', "
                            f"{melhor_score_linha:.0%}) foi atribuída a outro PDF da seção, com "
                            f"correspondência ainda melhor com aquele item.")
            else:
                detalhes = (f"Melhor similaridade encontrada: {melhor_score_linha:.0%}. "
                            f"Abaixo do limiar ({LIMIAR_TITULO:.0%}).")
            resultados.append({
                "seq":          pdf.stem,
                "titulo":       "Documento não identificado no Lattes",
                "complemento":  "",
                "ano":          "",
                "status":       "REPROVADO",
                "detalhes":     detalhes,
                "pdf_nome":     pdf.name,
                "pontos":       0,
                "subcategoria": "—",
                "tipo_pontos":  tipo_pontos,
            })

    return resultados  # pool já ordenado por nome → resultados em ordem dos PDFs


# ── Estilos Excel ─────────────────────────────────────────────────────────────

def hex_fill(h):
    return PatternFill("solid", start_color=h, fgColor=h)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Geração do relatório Excel (por candidato) ────────────────────────────────

# ── Aba "Declarado": autodeclaração do candidato (Anexo II) ──────────────────

_ANEXO2_SECAO_LABELS = {
    1: "1. Artigos publicados em periódicos",
    2: "2. Resumos publicados em anais",
    3: "3. Livros publicados",
    4: "4. Organização de livro",
    5: "5. Capítulo de livro",
    6: "6. Participação em bancas de TCC",
    7: "7. Atuação docente — ensino superior",
    8: "8. Atuação docente — ensino básico",
    9: "9. Atuação profissional não docente",
    10: "10. Participação em projeto de pesquisa",
    11: "11. Participação em projeto de extensão",
    12: "12. Participação em projeto de ensino",
    13: "13. Comissão organizadora de evento científico",
}


def _achar_anexo2(pasta: Path) -> Path | None:
    """Localiza, na pasta do candidato, o PDF de autodeclaração — geralmente
    um arquivo com nome parecido com "Anexo-II-Pontos-no-currículo" (varia
    de candidato pra candidato: maiúsculas, hífen/espaço, sufixo ".docx.pdf",
    ordem das palavras — ex. "Pontos Currículo - Fulano" etc.).

    Casa pelo nome normalizado primeiro (rápido); se não achar, confere o
    CONTEÚDO de cada PDF da pasta em busca do cabeçalho característico do
    Anexo II — cobre o caso de o nome do arquivo vir com encoding
    corrompido (zip extraído errado troca "í"/"é" por caracteres tipo
    "Ý"/"Ú", o que não é "acento a menos" e sobrevive à normalização de
    acentos comum)."""
    candidatos = list(pasta.glob("*.pdf")) + list(pasta.glob("*.PDF"))
    for pdf in candidatos:
        nome_norm = re.sub(r"[^a-z0-9]", "", normalizar(pdf.stem))
        if "anexoii" in nome_norm and ("pontos" in nome_norm or "curriculo" in nome_norm):
            return pdf
    for pdf in candidatos:
        texto = extrair_texto_pdf(pdf, max_paginas=2)
        if texto and not texto.startswith("__ERRO_PDF__") and "anexo ii" in normalizar(texto[:600]):
            return pdf
    return None


def _extrair_declarado_anexo2(pasta: Path) -> dict | None:
    """Lê o Anexo II autodeclarado pelo candidato e retorna a pontuação que
    ELE declarou, por seção — pra comparação manual com o que o
    verificador aprovou a partir dos comprovantes. Retorna None se não
    achar o arquivo ou não conseguir interpretar a tabela.

    O layout do PDF (gerado a partir de um .docx) quebra linha de forma
    inconsistente — quantidade e pontos às vezes ficam na mesma linha do
    quesito, às vezes em linhas seguintes separadas por texto ("por
    mês"). Por isso o parser não tenta casar "quesito → valores" numa
    linha só: isola o bloco de texto de cada seção numerada (do "N." até
    o próximo "N+1.") e pega o ÚLTIMO número inteiro e o ÚLTIMO número
    decimal do bloco como quantidade/pontos — robusto à quebra de linha,
    mas assume que não há mais de um par quantidade/pontos por seção
    (verdade pra todas as seções exceto Artigos, que tem sub-total
    próprio tratado à parte).
    """
    arquivo = _achar_anexo2(pasta)
    if arquivo is None:
        return None

    texto = extrair_texto_pdf(arquivo, max_paginas=10)
    if not texto.strip() or texto.startswith("__ERRO_PDF__"):
        return None

    m_total = re.search(r"Total de pontos:?\s*_*([\d.,]+)_*", texto)
    total_declarado = float(m_total.group(1).replace(",", ".")) if m_total else None

    headers = list(re.finditer(r"(?:^|\n)(\d{1,2})\.\s*(?=[A-ZÀ-Ú])", texto))
    marcos  = [(int(h.group(1)), h.start(), h.end()) for h in headers]

    secoes: dict[int, tuple[int | None, float]] = {}
    for i, (sec, _blk_start, cont_start) in enumerate(marcos):
        blk_end = marcos[i + 1][1] if i + 1 < len(marcos) else len(texto)
        bloco   = texto[cont_start:blk_end]

        if sec == 1:
            m = re.search(r"Sub-total\s+(\d+)\s+([\d.,]+)", bloco)
            if m:
                secoes[sec] = (int(m.group(1)), float(m.group(2).replace(",", ".")))
                continue
            # "Sub-total" às vezes vem em branco no PDF mesmo com as linhas
            # de Qualis individuais preenchidas (Primeiro/Demais por
            # nível) — soma o par quantidade/pontos de cada linha que tem
            # valor, como fallback.
            pares_linha = re.findall(r"\b(\d+)\s+(\d+[.,]\d+)\b", bloco)
            if pares_linha:
                qtd_total = sum(int(q) for q, _ in pares_linha)
                pts_total = sum(float(p.replace(",", ".")) for _, p in pares_linha)
                secoes[sec] = (qtd_total, pts_total)
            continue

        decimais = re.findall(r"\d+[.,]\d+", bloco)
        if not decimais:
            continue  # seção sem valor declarado (candidato não pontuou aí)
        # Prefere um par ADJACENTE "quantidade pontos" (mesmo formato do
        # fallback de artigo acima) — texto livre depois da linha (título
        # de cada item, às vezes com ano dentro, ex.: "CAMINHADA NÓRDICA
        # PARA IDOSOS 2023") pode ter número solto que não é a quantidade
        # de verdade; pegar só "o último inteiro do bloco inteiro" cai
        # nessa pegadinha. Só cai pro inteiro solto se não achar par.
        par = re.search(r"\b(\d+)\s+(\d+[.,]\d+)\b", bloco)
        if par:
            secoes[sec] = (int(par.group(1)), float(par.group(2).replace(",", ".")))
        else:
            inteiros = re.findall(r"\b\d+\b", re.sub(r"\d+[.,]\d+", " ", bloco))
            qtd = int(inteiros[-1]) if inteiros else None
            secoes[sec] = (qtd, float(decimais[-1].replace(",", ".")))

    if not secoes and total_declarado is None:
        return None
    return {"arquivo": arquivo.name, "total": total_declarado, "secoes": secoes}


# Seção do Anexo II → tipo(s) interno(s) correspondente(s) em `resultados`,
# usado só pra agregar a aba Resumo por seção do Anexo II (mesma numeração
# do criterios.xlsx/declaração). Atuação (7/8/9) não está aqui porque vem de
# `atuacoes` (lista separada), agregada por "categoria" — ver
# gerar_relatorio. Seção 12 (projeto de ensino) não tem tipo interno: o
# sistema não extrai essa categoria do Lattes, então nunca aparece do lado
# "Sistema" — isso é honesto (não é um bug, é uma categoria não implementada).
_ANEXO2_TIPO_MAP = {
    1:  {"artigo"},
    2:  {"trabalho_completo", "resumo"},
    3:  {"livro_publicado"},
    4:  {"livro_organizado"},
    5:  {"capitulo"},
    6:  {"banca"},
    10: {"projeto_pesquisa"},
    11: {"projeto_extensao"},
    13: {"organizacao_evento"},
}
_ANEXO2_SECAO_ATUACAO = {7: "ensino_superior", 8: "ensino_basico", 9: "nao_docencia"}

# Seções pontuadas "por mês" no Anexo II — Decl. Qtd. já vem em meses pra
# essas (é o que o candidato declarou), então Sist. Aprov. precisa somar
# meses dos itens aprovados em vez de contar itens, senão a comparação
# mistura unidades diferentes (itens x meses).
_ANEXO2_SECOES_MENSAL = {7, 8, 9, 10, 11, 12}


def _pontos_por_secao(resultados: list[dict], atuacoes: list[dict]) -> dict[int, float]:
    """Soma os pontos aprovados por seção do Anexo II (1 a 13) — mesmo
    agrupamento usado na aba "Resumo" do relatório individual (ver
    gerar_relatorio), só que aqui pra alimentar o ranking geral quesito a
    quesito em vez do agregado produção/atuação."""
    pontos: dict[int, float] = {}
    for sec in _ANEXO2_SECAO_LABELS:
        if sec in _ANEXO2_SECAO_ATUACAO:
            categoria = _ANEXO2_SECAO_ATUACAO[sec]
            pts = sum(a["pontos"] for a in atuacoes
                      if a.get("categoria") == categoria and a.get("status") == "APROVADO")
        else:
            tipos = _ANEXO2_TIPO_MAP.get(sec, set())
            pts = sum(r["pontos"] for r in resultados
                      if r.get("tipo_pontos") in tipos and r.get("status") == "APROVADO")
        pontos[sec] = round(pts, 3)
    return pontos


def gerar_relatorio(resultados: list[dict], nome_pesquisador: str,
                    caminho_saida: Path, atuacoes: list[dict] | None = None,
                    pasta_candidato: Path | None = None):
    wb = Workbook()

    # ── Aba 1: Relatório detalhado ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Relatório"
    for col, w in zip("ABCDEFGHI", [10, 42, 28, 8, 14, 10, 10, 44, 20]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Relatório de Verificação de Comprovantes — {nome_pesquisador}"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill("1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Arial", italic=True, size=9, color="5F5E5A")
    c.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 16

    # "Classificação" = subcategoria determinada por calcular_pontos_item —
    # pro artigo é o estrato Qualis (A1, A2, B1...); pras demais categorias
    # é a subcategoria equivalente (Nacional/Internacional, Coordenador/
    # Integrante, etc.), já calculada mas até então não exibida na planilha.
    for col, h in enumerate(["Código", "Título / Descrição", "Info. complementar",
                              "Ano / Autoria", "Classificação", "Status", "Pontos",
                              "Detalhes da verificação", "PDF encontrado"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hex_fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[3].height = 22

    secao_labels = {
        "1": "1. Artigos publicados em periódicos",
        "2": "2. Resumos/trabalhos publicados em anais de congressos",
        "4": "4. Livros publicados",
        "5": "5. Capítulos de livros",
        "6": "6. Participação em bancas de TCC",
        "10": "10. Projetos de pesquisa",
        "11": "11. Projetos de extensão",
        "13": "13. Organização de evento",
        "14": "14. Iniciação Científica (participação)",
        "16": "16. Produção Artística/Cultural — Internacional",
        "17": "17. Produção Artística/Cultural — Nacional",
        "18": "18. Produção Artística/Cultural — Regional/Local",
    }
    secao_atual, row = "", 4

    for res in resultados:
        secao = res["seq"].split(".")[0]
        if secao != secao_atual:
            secao_atual = secao
            ws.merge_cells(f"A{row}:I{row}")
            c = ws.cell(row=row, column=1, value=secao_labels.get(secao, f"Seção {secao}"))
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = hex_fill("2E75B6")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin_border()
            ws.row_dimensions[row].height = 22
            row += 1

        status = res["status"]
        status_cor, bg_status = {
            "APROVADO":  ("27500A", "EAF3DE"),
            "REPROVADO": ("791F1F", "FCEBEB"),
            "SEM PDF":   ("633806", "FAEEDA"),
            "ERRO PDF":  ("444441", "F5F5F3"),
            "SEM TEXTO": ("444441", "F5F5F3"),
        }.get(status, ("444441", "F5F5F3"))

        try:
            bg = "FFFFFF" if int(res["seq"].split(".")[1]) % 2 == 0 else "F7FBFF"
        except (IndexError, ValueError):
            bg = "FFFFFF"

        def cel(col, val, bold=False, align="left", bg_ov=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=9, bold=bold, color="2C2C2A")
            c.fill = hex_fill(bg_ov or bg)
            c.alignment = Alignment(horizontal=align, vertical="top",
                                    wrap_text=True, indent=1 if align == "left" else 0)
            c.border = thin_border()

        cel(1, res["seq"], bold=True, align="center")
        cel(2, res["titulo"])
        cel(3, res.get("complemento", ""))
        cel(4, res.get("ano", ""), align="center")
        cel(5, res.get("subcategoria", ""), align="center")

        c2 = ws.cell(row=row, column=6, value=status)
        c2.font = Font(name="Arial", bold=True, size=9, color=status_cor)
        c2.fill = hex_fill(bg_status)
        c2.alignment = Alignment(horizontal="center", vertical="top")
        c2.border = thin_border()

        pts = res.get("pontos", 0) if status == "APROVADO" else 0
        c3 = ws.cell(row=row, column=7, value=pts if pts else "")
        c3.font = Font(name="Arial", size=9, bold=True,
                       color="27500A" if pts else "AAAAAA")
        c3.fill = hex_fill(bg)
        c3.alignment = Alignment(horizontal="center", vertical="top")
        c3.border = thin_border()

        cel(8, res.get("detalhes", ""))
        cel(9, res.get("pdf_nome", "—"), align="center")
        ws.row_dimensions[row].height = 32
        row += 1

    # Atuação profissional entra na MESMA aba Relatório, não numa aba à
    # parte — candidato cuja produção é só atuação (ex.: sem nenhum
    # artigo/projeto) ficaria sem NENHUM detalhe item a item salvo no
    # arquivo, já que ela é uma lista separada de `resultados`. Agrupa por
    # CATEGORIA (não pelo prefixo numérico do "seq") porque esse prefixo
    # colide com o de outros tipos internos (banca também usa "7").
    _ORDEM_CATEGORIA_ATUACAO = ["ensino_superior", "ensino_basico", "nao_docencia"]
    atuacoes_ordenadas = sorted(
        atuacoes or [],
        key=lambda a: _ORDEM_CATEGORIA_ATUACAO.index(a.get("categoria", "nao_docencia"))
                      if a.get("categoria") in _ORDEM_CATEGORIA_ATUACAO else 99,
    )
    categoria_atual = None
    for at in atuacoes_ordenadas:
        categoria = at.get("categoria", "nao_docencia")
        if categoria != categoria_atual:
            categoria_atual = categoria
            secao_num = _SECAO_ATUACAO.get(categoria, "9")
            ws.merge_cells(f"A{row}:I{row}")
            c = ws.cell(row=row, column=1,
                        value=f"{secao_num}. Atuação Profissional — {LABEL_ATUACAO.get(categoria, categoria)}")
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = hex_fill("2E75B6")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin_border()
            ws.row_dimensions[row].height = 22
            row += 1

        status = at.get("status", "SEM PDF")
        status_cor, bg_status = {
            "APROVADO":  ("27500A", "EAF3DE"),
            "REPROVADO": ("791F1F", "FCEBEB"),
            "SEM PDF":   ("633806", "FAEEDA"),
            "ERRO PDF":  ("444441", "F5F5F3"),
        }.get(status, ("444441", "F5F5F3"))

        try:
            bg = "FFFFFF" if int(at.get("seq", "0.0").split(".")[1]) % 2 == 0 else "F7FBFF"
        except (IndexError, ValueError):
            bg = "FFFFFF"

        def cel_at(col, val, bold=False, align="left", bg_ov=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=9, bold=bold, color="2C2C2A")
            c.fill = hex_fill(bg_ov or bg)
            c.alignment = Alignment(horizontal=align, vertical="top",
                                    wrap_text=True, indent=1 if align == "left" else 0)
            c.border = thin_border()

        meses = at.get("meses", 0)
        cel_at(1, at.get("seq", ""), bold=True, align="center")
        cel_at(2, at.get("instituicao", ""))
        cel_at(3, at.get("descricao", ""))
        cel_at(4, at.get("periodo", ""), align="center")
        cel_at(5, f"{meses}m" if meses else "", align="center")

        c2 = ws.cell(row=row, column=6, value=status)
        c2.font = Font(name="Arial", bold=True, size=9, color=status_cor)
        c2.fill = hex_fill(bg_status)
        c2.alignment = Alignment(horizontal="center", vertical="top")
        c2.border = thin_border()

        pts = at.get("pontos", 0) if status == "APROVADO" else 0
        c3 = ws.cell(row=row, column=7, value=pts if pts else "")
        c3.font = Font(name="Arial", size=9, bold=True, color="27500A" if pts else "AAAAAA")
        c3.fill = hex_fill(bg)
        c3.alignment = Alignment(horizontal="center", vertical="top")
        c3.border = thin_border()

        cel_at(8, at.get("detalhes_pdf", ""))
        cel_at(9, at.get("pdf_nome", "—"), align="center")
        ws.row_dimensions[row].height = 32
        row += 1

    # ── Aba 2: Resumo ─────────────────────────────────────────────────────────
    # Autodeclaração do candidato (Anexo II), se o arquivo estiver na pasta —
    # calculada aqui porque a aba Resumo passa a mostrar declarado x sistema
    # lado a lado, seção a seção, em vez de uma aba separada.
    declarado = _extrair_declarado_anexo2(pasta_candidato) if pasta_candidato is not None else None
    decl_secoes = declarado["secoes"] if declarado else {}

    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 40
    for col in "BCDEF":
        ws2.column_dimensions[col].width = 15

    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = "Resumo por Categoria — Declarado (Anexo II) x Sistema"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill("1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    if declarado is not None:
        ws2.merge_cells("A2:F2")
        c = ws2["A2"]
        c.value = f"Declarado conforme: {declarado['arquivo']}"
        c.font = Font(name="Arial", italic=True, size=9, color="5F5E5A")
        c.alignment = Alignment(horizontal="right")

    for col, h in enumerate(["Categoria", "Decl. Qtd.", "Decl. Pts",
                              "Sist. Aprov.", "Sist. Pts",
                              "Diferença (Pts)"], 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hex_fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws2.row_dimensions[3].height = 26

    r2 = 4
    tot_decl_pts = tot_sist_pts = 0
    for sec, label in _ANEXO2_SECAO_LABELS.items():
        if sec in _ANEXO2_SECAO_ATUACAO:
            categoria = _ANEXO2_SECAO_ATUACAO[sec]
            grupo_sistema = [a for a in (atuacoes or []) if a.get("categoria") == categoria]
        else:
            tipos = _ANEXO2_TIPO_MAP.get(sec, set())
            grupo_sistema = [x for x in resultados if x.get("tipo_pontos") in tipos]

        decl_qtd, decl_pts = decl_secoes.get(sec, (None, None))
        pts = sum(x.get("pontos", 0) for x in grupo_sistema if x.get("status") == "APROVADO")

        # Atuação e projetos são pontuados por MÊS (ver Decl. Qtd., que pro
        # Anexo II já vem em meses nessas seções) — "Sist. Aprov." precisa
        # mostrar a mesma unidade (soma de meses dos itens aprovados), não a
        # contagem de itens, senão a comparação não é como-com-como.
        if sec in _ANEXO2_SECOES_MENSAL:
            aprov = sum(x.get("meses") or 0 for x in grupo_sistema if x.get("status") == "APROVADO")
        else:
            aprov = sum(1 for x in grupo_sistema if x.get("status") == "APROVADO")

        if not grupo_sistema and decl_pts is None:
            continue  # nada declarado, nada encontrado — não polui a aba

        tot_decl_pts += decl_pts or 0
        tot_sist_pts += pts
        dif = round(pts - (decl_pts or 0), 3) if decl_pts is not None else None

        bg = "F7FBFF" if r2 % 2 == 0 else "FFFFFF"
        valores = [label, decl_qtd if decl_qtd is not None else "",
                   decl_pts if decl_pts is not None else "",
                   aprov, round(pts, 3), dif if dif is not None else ""]
        for col, val in enumerate(valores, 1):
            c = ws2.cell(row=r2, column=col, value=val)
            cor = "2C2C2A"
            if col == 4 and aprov:
                cor = "27500A"
            elif col == 6 and dif is not None:
                cor = "791F1F" if dif < 0 else ("27500A" if dif > 0 else "2C2C2A")
            c.font = Font(name="Arial", size=10, bold=(col == 6), color=cor)
            c.fill = hex_fill(bg)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                    vertical="center", indent=1 if col == 1 else 0)
            c.border = thin_border()
        ws2.row_dimensions[r2].height = 20
        r2 += 1

    # Decl. Qtd./Sist. Aprov. não são somadas no total — as seções misturam
    # unidades diferentes (itens em umas, meses em outras; ver
    # _ANEXO2_SECOES_MENSAL), então a soma bruta não significaria nada. Só
    # os pontos (mesma unidade em todas as seções) somam de verdade.
    tot_dif = round(tot_sist_pts - tot_decl_pts, 3)
    for col, val in enumerate(["TOTAL", "", round(tot_decl_pts, 3),
                                "", round(tot_sist_pts, 3),
                                tot_dif], 1):
        c = ws2.cell(row=r2, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = hex_fill("1F4E79")
        c.alignment = Alignment(horizontal="center" if col > 1 else "right",
                                vertical="center", indent=1 if col == 1 else 0)
        c.border = thin_border()
    ws2.row_dimensions[r2].height = 24

    if declarado is not None:
        r2 += 2
        ws2.merge_cells(f"A{r2}:F{r2}")
        nota = ws2.cell(row=r2, column=1, value=(
            "Nota: \"Decl.\" vem da autodeclaração do candidato no Anexo II — "
            "não necessariamente o que os comprovantes sustentam. Nas seções "
            "de atuação profissional e projetos, \"Qtd.\"/\"Aprov.\" estão em "
            "meses, não em número de itens. Categorias sem tipo "
            "correspondente no sistema atual (ex.: projeto de ensino) não "
            "têm coluna \"Sistema\" comparável e aparecem em branco ali. "
            "Para o detalhamento do que o sistema aprovou item a item, ver a "
            "aba \"Relatório\"."))
        nota.font = Font(name="Arial", italic=True, size=8, color="5F5E5A")
        nota.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r2].height = 40

    # Aba "Atuação Profissional" removida do relatório por candidato — o
    # detalhamento item a item continua no ranking geral (ranking.xlsx tem
    # uma coluna por quesito do Anexo II, ver _pontos_por_secao/gerar_ranking);
    # a aba Resumo acima já traz o total por categoria de atuação (seções
    # 7/8/9).

    wb.move_sheet("Resumo", offset=-1)  # Resumo antes de Relatório
    wb.save(caminho_saida)


# ── Ranking de candidatos ─────────────────────────────────────────────────────

def gerar_ranking(ranking: list[dict], caminho_saida: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Classificação"

    # Uma coluna por quesito do Anexo II (mesma numeração/rótulo de
    # _ANEXO2_SECAO_LABELS), na ordem oficial, mais Total Geral no final.
    secoes  = list(_ANEXO2_SECAO_LABELS.keys())
    n_cols  = 2 + len(secoes) + 1  # # + Candidato + quesitos + Total Geral
    col_tot = n_cols

    larguras = [5, 30] + [14] * len(secoes) + [16]
    for col, w in zip(range(1, n_cols + 1), larguras):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1)
    c.value = f"Classificação de Candidatos — gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill("1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = (["#", "Candidato"]
               + [_ANEXO2_SECAO_LABELS[s] for s in secoes]
               + ["Total Geral"])
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hex_fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 44

    ranking_sorted = sorted(ranking, key=lambda x: x["total"], reverse=True)

    for pos, cand in enumerate(ranking_sorted, 1):
        row = pos + 2
        bg  = "F7FBFF" if pos % 2 == 0 else "FFFFFF"
        cor_pos = "B8860B" if pos == 1 else ("888888" if pos == 2 else
                  "CD7F32" if pos == 3 else "2C2C2A")

        por_secao = cand.get("por_secao", {})
        vals = ([pos, cand["nome"]]
                + [por_secao.get(s, 0) for s in secoes]
                + [cand["total"]])
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            bold = col in (1, col_tot)
            c.font = Font(name="Arial", size=10, bold=bold,
                          color=cor_pos if col in (1, 2, col_tot) else "2C2C2A")
            c.fill = hex_fill(bg)
            c.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                    vertical="center", indent=1 if col == 2 else 0)
            c.border = thin_border()
        ws.row_dimensions[row].height = 22

    wb.save(caminho_saida)
    print(f"Ranking gerado: {caminho_saida}")


# ── Orquestrador principal ────────────────────────────────────────────────────

def verificar_curriculo(pasta_curriculo: str | Path,
                        criterios: dict | None = None,
                        qualis: dict | None = None,
                        campos_config: dict | None = None,
                        pasta_saida: Path | None = None) -> dict:
    """Processa um candidato e retorna dict com totais para o ranking."""
    pasta    = Path(pasta_curriculo)
    xml_path = _achar_lattes_xml(pasta)
    if xml_path is None:
        raise FileNotFoundError(f"lattes.xml não encontrado em {pasta}")

    if criterios is None:
        criterios, campos_config = carregar_criterios()
    elif campos_config is None:
        campos_config = _campos_config_padrao()
    if qualis is None:
        qualis = carregar_qualis()

    root = parse_xml(xml_path)
    dg   = root.find("DADOS-GERAIS")
    nome = dg.get("NOME-COMPLETO", pasta.name) if dg is not None else pasta.name

    print(f"\n{'='*60}")
    print(f"Verificando currículo: {nome}")
    print(f"{'='*60}")

    extratores = [
        (extrair_artigos,             "periodico",    "artigo"),
        (extrair_trabalhos_completos, "evento",       "trabalho_completo"),
        (extrair_resumos,             "evento",       "resumo"),
        (extrair_livros,              "editora",      "livro_publicado"),
        (extrair_capitulos,           "livro",        "capitulo"),
        (extrair_orientacoes,         "orientando",   "orientacao_mestrado"),
        (extrair_bancas,              "candidato",    "banca"),
        (extrair_organizacoes_evento, "evento",       "organizacao_evento"),
    ]

    resultados = []

    for extrator, campo_compl, tipo_pontos in extratores:
        # Categoria que a planilha de critérios atual não pontua (ex.:
        # orientação, nesse edital) nem vale buscar PDF pra ela — ganho
        # duplo: menos trabalho à toa, e evita colisão de numeração de
        # seção com outra categoria que USA aquele prefixo internamente
        # (ex.: orientação e banca dividiam o prefixo "6" antes desse
        # check existir).
        if not criterios.get(tipo_pontos):
            continue
        # extrair_artigos precisa do nome do pesquisador pra achar sua
        # própria ORDEM-DE-AUTORIA na lista de AUTORES (ver _ordem_autoria).
        items = extrator(root, nome) if tipo_pontos in ("artigo", "capitulo") else extrator(root)
        if not items:
            continue
        secao  = items[0]["seq"].split(".")[0]
        pool   = _pool_pdfs_secao(pasta, secao)
        if not pool:
            continue  # sem PDFs nesta seção → omite do relatório
        campos_por_sub = campos_config.get(tipo_pontos, {})
        print(f"\nSeção {secao}: {len(pool)} PDF(s) / {len(items)} item(s) no Lattes")

        res_secao = _processar_secao_por_pdf(
            items, pool, campo_compl, tipo_pontos, campos_por_sub, nome, criterios, qualis)
        for r in res_secao:
            simbolo = "✓" if r["status"] == "APROVADO" else "✗"
            print(f"  [{simbolo}] {r['pdf_nome']} → {r['seq']} — {r['titulo'][:50]}...")
        resultados.extend(res_secao)

    # Iniciação Científica — comprovação por PDF (sem XML)
    ic_items = extrair_ic_manual(pasta)
    if ic_items:
        print(f"\nSeção 14: {len(ic_items)} item(s)")
        campos_ic = _resolver_campos(campos_config.get("ic_manual"), "unidade")
        for item in ic_items:
            pdf_path = item["_pdf_path"]
            pontos_item, sub_item = calcular_pontos_item(item, "ic_manual", criterios, qualis)
            texto = extrair_texto_pdf(pdf_path)
            if texto.startswith("__ERRO_PDF__"):
                status, detalhes, pdf_nome = "ERRO PDF", texto, pdf_path.name
                pontos_item = 0
            else:
                res    = verificar_por_config(item, texto, campos_ic, nome)
                status = "APROVADO" if res["aprovado"] else "REPROVADO"
                detalhes, pdf_nome = res["detalhes"], pdf_path.name
            simbolo = "✓" if status == "APROVADO" else "✗"
            print(f"  [{simbolo}] {item['seq']} — Iniciação Científica")
            resultados.append({
                "seq":          item["seq"],
                "titulo":       item["titulo"],
                "complemento":  "",
                "ano":          "",
                "status":       status,
                "detalhes":     detalhes,
                "pdf_nome":     pdf_nome,
                "pontos":       pontos_item,
                "subcategoria": sub_item,
            })

    # Banca de especialização — comprovação por PDF (sem XML, mesmo padrão da IC)
    banca_esp_items = extrair_banca_especializacao_manual(pasta)
    if banca_esp_items:
        print(f"\nSeção 15: {len(banca_esp_items)} item(s)")
        campos_banca_esp = _resolver_campos(campos_config.get("banca"), "Especializacao")
        for item in banca_esp_items:
            pdf_path = item["_pdf_path"]
            pontos_item, sub_item = calcular_pontos_item(
                item, "banca_especializacao_manual", criterios, qualis)
            texto = extrair_texto_pdf(pdf_path)
            if texto.startswith("__ERRO_PDF__"):
                status, detalhes, pdf_nome = "ERRO PDF", texto, pdf_path.name
                pontos_item = 0
            else:
                res    = verificar_por_config(item, texto, campos_banca_esp, nome)
                status = "APROVADO" if res["aprovado"] else "REPROVADO"
                detalhes, pdf_nome = res["detalhes"], pdf_path.name
            simbolo = "✓" if status == "APROVADO" else "✗"
            print(f"  [{simbolo}] {item['seq']} — Banca de Especialização")
            resultados.append({
                "seq":          item["seq"],
                "titulo":       item["titulo"],
                "complemento":  "",
                "ano":          "",
                "status":       status,
                "detalhes":     detalhes,
                "pdf_nome":     pdf_nome,
                "pontos":       pontos_item,
                "subcategoria": sub_item,
            })

    # Produção Artística/Cultural — extrai do XML e verifica título no PDF
    prod_art_int, prod_art_nac, prod_art_reg = extrair_prod_artistica_xml(root)
    for tipo_art, lista_art, label_art in [
        ("prod_artistica_int", prod_art_int, "Internacional"),
        ("prod_artistica_nac", prod_art_nac, "Nacional"),
        ("prod_artistica_reg", prod_art_reg, "Regional/Local"),
    ]:
        if not lista_art:
            continue
        secao_num = lista_art[0]["seq"].split(".")[0]
        print(f"\nSeção {secao_num} (Prod. Artística — {label_art}): {len(lista_art)} item(s)")
        campos_art = campos_config.get(tipo_art, {})
        pool_art   = _pool_pdfs_secao(pasta, secao_num)
        if not pool_art:
            continue
        res_art = _processar_secao_por_pdf(
            lista_art, pool_art, "tipo_arte", tipo_art, campos_art, nome, criterios, qualis)
        for r in res_art:
            simbolo = "✓" if r["status"] == "APROVADO" else "✗"
            print(f"  [{simbolo}] {r['pdf_nome']} → {r['seq']} — {r['titulo'][:50]}...")
        resultados.extend(res_art)

    # Projetos de pesquisa e extensão
    proj_pesquisa, proj_extensao = extrair_projetos(root, nome)
    for tipo_proj, lista_proj in [
        ("projeto_pesquisa", proj_pesquisa),
        ("projeto_extensao", proj_extensao),
    ]:
        if not lista_proj:
            continue
        secao_num = lista_proj[0]["seq"].split(".")[0]
        print(f"\nSeção {secao_num}: {len(lista_proj)} item(s)")
        campos_proj = campos_config.get(tipo_proj, {})
        pool_proj   = _pool_pdfs_secao(pasta, secao_num)
        if not pool_proj:
            continue
        res_proj = _processar_secao_por_pdf(
            lista_proj, pool_proj, "papel", tipo_proj, campos_proj, nome, criterios, qualis)
        for r in res_proj:
            simbolo = "✓" if r["status"] == "APROVADO" else "✗"
            print(f"  [{simbolo}] {r['pdf_nome']} → {r['seq']} — {r['titulo'][:50]}...")
        resultados.extend(res_proj)

    # Atuação profissional — PDF-centric. Três categorias com pontuação
    # própria (docente ensino superior/básico, não docente) — cada uma tem
    # sua própria seção de PDFs na pasta do candidato (7/8/9, ver
    # _SECAO_ATUACAO), então o pool precisa juntar as três em vez de olhar
    # só uma seção fixa.
    print("\n── Atuação Profissional ──")
    pontos_atu   = criterios.get("atuacao", {})
    atuacoes_xml = extrair_atuacao_profissional(root, pontos_atu)
    pool_atu     = [pdf_texto for secao in _SECAO_ATUACAO.values()
                              for pdf_texto in _pool_pdfs_secao(pasta, secao)]
    atuacoes: list[dict] = []

    pool_atu_validos = []
    for pdf, texto in pool_atu:
        if not texto.strip() or texto.startswith("__ERRO_PDF__"):
            atuacoes.append({
                "seq": pdf.stem, "instituicao": "—", "descricao": "—",
                "periodo": "—", "ano_inicio": "", "categoria": "nao_docencia",
                "meses": 0, "pontos_unit": 0, "pontos": 0,
                "status": "ERRO PDF",
                "detalhes_pdf": "Falha na leitura do PDF.",
                "pdf_nome": pdf.name,
            })
            print(f"  [!] {pdf.name} — ERRO PDF")
        else:
            pool_atu_validos.append((pdf, texto))

    # Casamento é por INSTITUIÇÃO (nome + sigla), não por vínculo individual
    # — uma carteira de trabalho normalmente prova o contrato inteiro com o
    # empregador, promoções incluídas (ver _agrupar_atuacao_por_instituicao/
    # _verificar_atuacao_instituicao), não uma fase específica. A atribuição
    # ótima continua garantindo que cada PDF vá pra instituição com que
    # melhor combina, quando há mais de uma candidata.
    grupos_atu = _agrupar_atuacao_por_instituicao(atuacoes_xml)
    resultados_verif = [[_verificar_atuacao_instituicao(g, texto, nome) for g in grupos_atu]
                         for _, texto in pool_atu_validos]
    matriz_atu     = [[r["score"] if r["aprovado"] else -1.0 for r in linha] for linha in resultados_verif]
    atribuicao_atu = _atribuicao_otima(matriz_atu, 0.0)

    for i, (pdf, texto) in enumerate(pool_atu_validos):
        grupo_idx, _ = atribuicao_atu[i]

        if grupo_idx is not None:
            grupo = grupos_atu[grupo_idx]
            res   = resultados_verif[i][grupo_idx]

            # Meses = período TOTAL achado no PDF (data mais antiga à mais
            # nova, ver _periodo_do_pdf) — não a soma dos vínculos do
            # Lattes, que podem ter sido cadastrados com datas um pouco
            # diferentes das do documento oficial. Sem data reconhecível no
            # PDF, cai pra soma dos meses de cada vínculo do Lattes daquela
            # instituição (fallback por ano, ver extrair_atuacao_profissional).
            periodo_pdf = _periodo_do_pdf(texto, nome)
            if periodo_pdf is not None:
                inicio, fim = periodo_pdf
                meses = _meses_entre(inicio, fim)
                periodo_label = f"{inicio.strftime('%m/%Y')} – {fim.strftime('%m/%Y')}"
            else:
                meses = sum(it.get("meses", 0) for it in grupo["itens"])
                periodo_label = "; ".join(it["periodo"] for it in grupo["itens"])

            pts_unit = pontos_atu.get(grupo["categoria"], 0)
            cargos   = ", ".join(sorted({it["descricao"] for it in grupo["itens"]
                                         if it["descricao"] not in ("", "—")}))

            atuacoes.append({
                "seq":         pdf.stem,
                "instituicao": grupo["instituicao"],
                "descricao":   cargos or "—",
                "periodo":     periodo_label,
                "ano_inicio":  "",
                "categoria":   grupo["categoria"],
                "meses":       meses,
                "pontos_unit": pts_unit,
                "pontos":      meses * pts_unit,
                "status":      "APROVADO",
                "detalhes_pdf": f"{res['detalhes']} | Meses (período completo do PDF): {meses}",
                "pdf_nome":    pdf.name,
            })
            print(f"  [✓] {pdf.name} → {grupo['instituicao'][:45]} ({meses}m, {len(grupo['itens'])} vínculo(s) do Lattes)")
        else:
            atuacoes.append({
                "seq": pdf.stem, "instituicao": "Documento não identificado no Lattes",
                "descricao": "—", "periodo": "—", "ano_inicio": "", "categoria": "nao_docencia",
                "meses": 0, "pontos_unit": 0, "pontos": 0,
                "status": "REPROVADO",
                "detalhes_pdf": "PDF não corresponde a nenhuma instituição no Lattes.",
                "pdf_nome": pdf.name,
            })
            print(f"  [✗] {pdf.name} — não identificado")

    # Totais (somente APROVADO conta)
    pts_producao = sum(r["pontos"] for r in resultados if r["status"] == "APROVADO")
    pts_atuacao  = sum(a["pontos"] for a in atuacoes   if a.get("status") == "APROVADO")
    total        = pts_producao + pts_atuacao
    aprovados    = sum(1 for r in resultados if r["status"] == "APROVADO")
    sem_pdf      = sum(1 for r in resultados if r["status"] == "SEM PDF")
    sem_pdf     += sum(1 for a in atuacoes if a.get("status") == "SEM PDF")
    por_secao    = _pontos_por_secao(resultados, atuacoes)

    nome_arquivo = re.sub(r"[^\w\s-]", "", nome).strip().replace(" ", "_")
    destino_rel  = pasta_saida if pasta_saida is not None else pasta
    saida = destino_rel / f"relatorio_{nome_arquivo}.xlsx"
    gerar_relatorio(resultados, nome, saida, atuacoes=atuacoes, pasta_candidato=pasta)

    print(f"\n{'='*60}")
    print(f"Relatório: {saida}")
    print(f"Pontos produção (aprovados): {pts_producao}")
    print(f"Pontos atuação  (aprovados): {pts_atuacao}")
    print(f"TOTAL: {total}")
    print(f"{'='*60}\n")

    return {
        "nome":         nome,
        "pts_producao": pts_producao,
        "pts_atuacao":  pts_atuacao,
        "total":        total,
        "aprovados":    aprovados,
        "sem_pdf":      sem_pdf,
        "relatorio":    saida.name,
        "por_secao":    por_secao,
    }


def classificar_candidatos(candidatos_dir: Path = CANDIDATOS_DIR,
                           pasta_saida: Path | None = None):
    """Processa todos os candidatos e gera planilha de ranking."""
    criterios, campos_config = carregar_criterios()
    qualis                   = carregar_qualis()

    pastas = sorted(p for p in candidatos_dir.iterdir()
                    if p.is_dir() and _achar_lattes_xml(p) is not None)
    if not pastas:
        print(f"Nenhum candidato encontrado em {candidatos_dir}")
        return None

    print(f"\nProcessando {len(pastas)} candidato(s)...\n")
    destino = (pasta_saida or candidatos_dir.parent)
    ranking = []
    for pasta in pastas:
        try:
            dados = verificar_curriculo(pasta, criterios, qualis, campos_config,
                                        pasta_saida=destino)
            ranking.append(dados)
        except Exception as e:
            print(f"[ERRO] {pasta.name}: {e}")
    saida_ranking = destino / "ranking.xlsx"
    gerar_ranking(ranking, saida_ranking)
    return saida_ranking


# ── Entrada ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) == 1:
        # Sem argumento: processa todos os candidatos
        classificar_candidatos()
    else:
        alvo = Path(sys.argv[1])
        if _achar_lattes_xml(alvo) is not None:
            # Candidato individual
            verificar_curriculo(alvo)
        elif alvo.is_dir():
            # Pasta de candidatos
            classificar_candidatos(alvo)
        else:
            print(f"Caminho inválido: {alvo}")
            sys.exit(1)
